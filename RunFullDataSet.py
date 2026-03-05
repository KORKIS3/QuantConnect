#!/usr/bin/env python3
"""
RunFullDataSet - iterate over all local YM_intraday CSVs and create interactive plots.

Usage:
    python RunFullDataSet.py

The script looks for files matching: YM_intraday_<YYYY-MM-DD>_0930-1000.csv
and calls the project's existing `get_ym_intraday` and `plot_intraday_data`.
"""

import glob
import os
from typing import List

import pandas as pd

from plotFigure import ChartPlotter


def _load_csv_as_df(fp: str) -> pd.DataFrame:
    # Robust CSV loader to support older formats:
    # - auto-detect delimiter
    # - detect 'datetime'/'timestamp' column or separate 'Date' + 'Time' columns
    # - parse index to timezone-aware US/Eastern
    # - coerce numeric columns and map common alternate names
    import pytz
    import re

    est = pytz.timezone('US/Eastern')

    # First attempt: read with automatic delimiter detection
    try:
        df_try = pd.read_csv(fp, sep=None, engine='python', thousands=',', skipinitialspace=True)
    except Exception:
        # fallback to simple read
        df_try = pd.read_csv(fp, thousands=',', skipinitialspace=True)

    # If first column looks like an index of datetimes, try to parse it directly
    df = df_try.copy()

    def _parse_datetime_column(df_local):
        # 1) look for obvious names
        candidates = []
        for cname in df_local.columns:
            low = cname.lower()
            if 'datetime' in low or 'timestamp' in low or 'date_time' in low or 'time' == low:
                candidates.append(cname)
            if 'date' in low and 'time' in low:
                candidates.append(cname)
        # 2) date + time pair
        col_names = [c.lower() for c in df_local.columns]
        if 'date' in col_names and 'time' in col_names:
            return pd.to_datetime(df_local.loc[:, df_local.columns[[i for i,c in enumerate(col_names) if c=='date'][0]]].astype(str) + ' ' +
                                   df_local.loc[:, df_local.columns[[i for i,c in enumerate(col_names) if c=='time'][0]]].astype(str), errors='coerce')

        # 3) try any single column that parses well as datetimes
        best_col = None
        best_nonnull = 0
        for c in df_local.columns:
            try:
                parsed = pd.to_datetime(df_local[c], errors='coerce')
                nonnull = parsed.notna().sum()
                if nonnull > best_nonnull:
                    best_nonnull = nonnull
                    best_col = c
            except Exception:
                continue

        if best_col and best_nonnull > 0:
            return pd.to_datetime(df_local[best_col], errors='coerce')

        return None

    dt_index = None
    # If index-like column exists (Unnamed: 0 or first column) and looks datetime-ish
    first_col = df.columns[0] if len(df.columns) > 0 else None
    if first_col is not None:
        # If the headerless index was used as an index in some files, try parsing it
        series_first = df[first_col]
        parsed_first = pd.to_datetime(series_first, errors='coerce')
        if parsed_first.notna().sum() > 0:
            dt_index = parsed_first

    if dt_index is None:
        dt_index = _parse_datetime_column(df)

    if dt_index is not None and dt_index.notna().sum() > 0:
        # set as index and drop original column(s) if needed
        # find which column matched; try to locate and drop only that column
        try:
            # if dt_index came from a combination, it will be a Series with same length
            df = df.copy()
            df.index = dt_index
        except Exception:
            pass

    # If index still not datetime, attempt a second read with index_col=0 parse_dates=True
    if not pd.api.types.is_datetime64_any_dtype(df.index):
        try:
            df2 = pd.read_csv(fp, index_col=0, parse_dates=True, thousands=',', skipinitialspace=True)
            if pd.api.types.is_datetime64_any_dtype(df2.index):
                df = df2
        except Exception:
            pass

    # Ensure timezone-awareness (US/Eastern)
    try:
        if df.index.tz is None:
            df.index = pd.to_datetime(df.index, errors='coerce')
            df.index = df.index.tz_localize(est, ambiguous='NaT', nonexistent='shift_forward')
        else:
            df.index = pd.to_datetime(df.index).tz_convert(est)
    except Exception:
        # if conversion fails, leave as-is
        pass

    # Ensure numeric types for price/volume columns and map common alternate names
    for col in ['Open', 'High', 'Low', 'Close', 'Volume']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')

    cols_lower = {c.lower(): c for c in df.columns}
    def find_col(key_substr):
        for k, orig in cols_lower.items():
            if key_substr in k:
                return orig
        return None

    mapping = {}
    for key in ['open', 'high', 'low', 'close', 'volume']:
        proper = key.capitalize()
        if proper not in df.columns:
            found = find_col(key)
            if found:
                mapping[found] = proper

    if mapping:
        df = df.rename(columns=mapping)

    # If still missing core columns, try position-based inference among numeric columns
    core = ['Open', 'High', 'Low', 'Close']
    missing = [c for c in core if c not in df.columns]
    if missing:
        numeric_cols = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]
        if len(numeric_cols) >= 4:
            df = df.rename(columns={numeric_cols[0]: 'Open', numeric_cols[1]: 'High', numeric_cols[2]: 'Low', numeric_cols[3]: 'Close'})

    return df


def _compute_trade_stats(plotter: ChartPlotter) -> dict:
    # Signals already populated by frame-stepping in run_from_desktop;
    # only fall back to detect_all_signals_once if nothing was detected
    if not plotter.state.detected_buy_signals and not plotter.state.detected_sell_signals:
        if not plotter.state.all_signals_detected:
            plotter.detect_all_signals_once()

    buys = {t: p for t, p in plotter.state.detected_buy_signals.items()}
    sells = {t: p for t, p in plotter.state.detected_sell_signals.items()}

    # Merge events and compute P/L per completed trade.
    # Additionally simulate minute-by-minute total P/L to record:
    # - whether total P/L ever reached >= 100 ('Captured 100 points')
    # - highest total P/L during the window ('p/l high')
    # - P/L at the moment the app performed a liquidation (first close with abs(pl) >= 100)
    import pytz

    est = pytz.timezone('US/Eastern')

    events = []
    for t, p in buys.items():
        events.append((t, 'buy', p))
    for t, p in sells.items():
        events.append((t, 'sell', p))
    events.sort(key=lambda x: x[0])

    # normalize event times to plotter.data index tz
    events_by_time = {}
    for t, typ, price in events:
        try:
            evt = pd.to_datetime(t)
            if evt.tzinfo is None:
                evt = evt.tz_localize(est)
            else:
                evt = evt.tz_convert(est)
        except Exception:
            evt = pd.to_datetime(t)
        events_by_time.setdefault(evt, []).append((typ, price))

    trades: List[dict] = []
    position = 'flat'
    entry_price = None
    entry_time = None

    cumulative_realized_pl = 0.0
    max_total_pl = 0.0
    captured100 = False
    liquidation_p_l = None
    liquidation_trade_pl = None

    # iterate through each timestamp in the data and process events at that time
    for ts in plotter.data.index:
        # process events at this timestamp (if any)
        evs = events_by_time.get(ts, [])
        for typ, price in evs:
            if typ == 'buy':
                if position == 'flat':
                    position = 'long'
                    entry_price = price
                    entry_time = ts
                elif position == 'short':
                    pl = entry_price - price
                    cumulative_realized_pl += pl
                    trades.append({'entry_time': entry_time, 'exit_time': ts, 'pl': pl})
                    # record liquidation if this close exceeded threshold
                    if pl >= 100 and liquidation_p_l is None:
                        liquidation_p_l = cumulative_realized_pl
                        liquidation_trade_pl = float(pl)
                    position = 'long'
                    entry_price = price
                    entry_time = ts
            else:  # sell
                if position == 'flat':
                    position = 'short'
                    entry_price = price
                    entry_time = ts
                elif position == 'long':
                    pl = price - entry_price
                    cumulative_realized_pl += pl
                    trades.append({'entry_time': entry_time, 'exit_time': ts, 'pl': pl})
                    if pl >= 100 and liquidation_p_l is None:
                        liquidation_p_l = cumulative_realized_pl
                        liquidation_trade_pl = float(pl)
                    position = 'short'
                    entry_price = price
                    entry_time = ts

        # compute current total P/L (realized + unrealized)
        try:
            current_close = float(plotter.data['Close'].loc[ts])
        except Exception:
            current_close = float(plotter.data['Close'].iloc[-1])

        unrealized = 0.0
        if position == 'long' and entry_price is not None:
            unrealized = current_close - entry_price
        elif position == 'short' and entry_price is not None:
            unrealized = entry_price - current_close

        total_pl = cumulative_realized_pl + unrealized
        if total_pl > max_total_pl:
            max_total_pl = total_pl
        if (not captured100) and total_pl >= 100.0:
            captured100 = True

    # at end of data, if still in a position, close at last price
    if position != 'flat' and entry_price is not None:
        last_price = float(plotter.data['Close'].iloc[-1])
        if position == 'long':
            pl = last_price - entry_price
        else:
            pl = entry_price - last_price
        cumulative_realized_pl += pl
        trades.append({'entry_time': entry_time, 'exit_time': plotter.data.index[-1], 'pl': pl})
        if pl >= 100 and liquidation_p_l is None:
            liquidation_p_l = cumulative_realized_pl
            liquidation_trade_pl = float(pl)

    total_pl = cumulative_realized_pl
    wins = sum(1 for t in trades if t['pl'] > 0)
    losses = sum(1 for t in trades if t['pl'] <= 0)
    total_trades = len(trades)
    win_pct = (wins / total_trades * 100) if total_trades > 0 else 0.0

    # mark captured if either total/unrealized hit >=100 OR any single trade had abs(pl) >= 100
    captured_by_trade = any(t['pl'] >= 100 for t in trades)
    captured_flag = captured100 or captured_by_trade

    return {
        'final_pl': total_pl,
        'winning_trades': wins,
        'losing_trades': losses,
        'win_pct': win_pct,
        'total_trades': total_trades,
        'captured_100': 'Yes' if captured_flag else 'No',
        'pl_high': float(max_total_pl),
        'liquidation_pl': float(liquidation_p_l) if liquidation_p_l is not None else None,
        'liquidation_trade_pl': float(liquidation_trade_pl) if liquidation_trade_pl is not None else None,
    }



def run_from_desktop(desktop_subfolder: str = "FebData", start_time: str = "09:30", end_time: str = "10:00"):
    desktop_dir = os.path.join(os.path.expanduser('~'), 'Desktop', desktop_subfolder)
    # Accept any CSV filenames in the folder; we'll inspect contents to find suitable intraday files
    pattern = os.path.join(desktop_dir, "*.csv")
    files = sorted(glob.glob(pattern))
    if not files:
        print(f"No files found in {desktop_dir} matching pattern")
        return

    results = []
    output_dir = os.path.join(os.path.expanduser('~'), 'Desktop', 'Trading', 'Temp')
    os.makedirs(output_dir, exist_ok=True)

    for fp in files:
        fname = os.path.basename(fp)
        print('\n' + '='*60)
        print(f"Processing file: {fname}")

        try:
            df = _load_csv_as_df(fp)
        except Exception as e:
            print(f"Failed to load {fp}: {e}")
            continue

        # Quick check: does this file contain the expected time window?
        try:
            times = df.index.time
            has_start = any(t.strftime('%H:%M') == start_time for t in times)
            has_end = any(t.strftime('%H:%M') == end_time for t in times)
            if not (has_start or has_end):
                print(f"Skipping {fname}: does not contain {start_time}–{end_time} timestamps")
                continue
        except Exception:
            print(f"Skipping {fname}: could not verify timestamps")
            continue

        # extract date from filename (support YYYY-MM-DD or YYYYMMDD).
        # If filename contains a month name + day (e.g. 'feb 3'), prefer that.
        import re, calendar, datetime

        target_date = ''
        m = re.search(r"(\d{4}-\d{2}-\d{2})", fname)
        if m:
            target_date = m.group(1)
        else:
            m2 = re.search(r"(\d{8})", fname)
            if m2:
                s = m2.group(1)
                target_date = f"{s[0:4]}-{s[4:6]}-{s[6:8]}"

        # Prefer month-name + day patterns (e.g. 'feb 3') when present in filename
        m3 = re.search(r"(?i)\b(jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)[a-z]*[\s_\-]*(\d{1,2})\b", fname)
        if m3:
            month_str = m3.group(1)
            day = int(m3.group(2))
            try:
                month_num = list(calendar.month_abbr).index(month_str[:3].title())
                # Use fixed year 2026 per user request
                year = 2026
                target_date = f"{year:04d}-{month_num:02d}-{day:02d}"
            except Exception:
                pass

        # fallback to the first timestamp in the dataframe
        if not target_date:
            try:
                target_date = df.index[0].strftime('%Y-%m-%d')
            except Exception:
                target_date = ''

        # Ensure we have numeric high/low data
        if 'High' not in df.columns or 'Low' not in df.columns:
            print(f"Skipping {fname}: missing High/Low columns after mapping")
            continue
        if df['High'].dropna().empty or df['Low'].dropna().empty:
            print(f"Skipping {fname}: High/Low columns contain no numeric data")
            continue

        # Create ChartPlotter and step through every frame exactly like Main.py does
        # interactively — this ensures update_signals_incremental is used (same code path)
        plotter = ChartPlotter(df, target_date, start_time, end_time, output_dir, batch_mode=True)
        try:
            plotter.create_figure()
            # Finalize layout before any aspect-ratio-dependent calculations
            plotter.fig.canvas.draw()
        except Exception:
            pass

        # Step through each frame to trigger update_signals_incremental per minute
        try:
            for frame in range(len(df)):
                plotter.update_plot(frame)
                if plotter.state.trading_halted:
                    break
        except Exception:
            pass

        # Save figure as JPEG into ~/Desktop/TradingPics/ using the CSV-derived date as filename
        try:
            home = os.path.expanduser('~')
            image_dir = os.path.join(home, 'Desktop', 'TradingPics')
            os.makedirs(image_dir, exist_ok=True)
            import re
            # prefer the extracted target_date for filename when available
            if target_date:
                base_name = target_date
            else:
                base = os.path.splitext(fname)[0]
                base_name = re.sub(r'[^A-Za-z0-9_.-]', '_', base)

            # ensure we don't overwrite an existing file (append suffix if needed)
            img_name = f"{base_name}.jpg"
            img_path = os.path.join(image_dir, img_name)
            suffix = 1
            while os.path.exists(img_path):
                img_name = f"{base_name}_{suffix}.jpg"
                img_path = os.path.join(image_dir, img_name)
                suffix += 1

            if hasattr(plotter, 'fig') and getattr(plotter, 'fig') is not None:
                try:
                    # Ensure the figure is fully rendered by updating to the final frame
                    try:
                        final_frame = max(0, len(getattr(plotter, 'data', [])) - 1)
                        plotter.state.current_frame = final_frame
                        plotter.update_plot(final_frame)
                        # draw and flush events to populate artists
                        try:
                            plotter.fig.canvas.draw()
                            plotter.fig.canvas.flush_events()
                        except Exception:
                            pass
                    except Exception:
                        pass

                    plotter.fig.savefig(img_path, dpi=150, bbox_inches='tight')
                    print(f"Saved plot image: {img_path}")
                except Exception as e:
                    print(f"Failed to save plot image for {fname}: {e}")
        except Exception as e:
            print('Failed to prepare image directory:', e)

        # Close figure to avoid GUI windows
        try:
            import matplotlib.pyplot as _plt
            _plt.close(plotter.fig)
        except Exception:
            pass

        stats = _compute_trade_stats(plotter)

        results.append({
            'FileName': fname,
            'Date': target_date,
            'final_P/L': stats['final_pl'],
            'winning_trades': stats['winning_trades'],
            'losing_trades': stats['losing_trades'],
            'win_pct': stats['win_pct'],
            'Captured 100 points': stats.get('captured_100', 'No'),
            'p/l high': stats.get('pl_high', 0.0),
            'p/l liquidation trade': stats.get('liquidation_trade_pl', None),
        })

    # Save results to Excel on Desktop
    out_xlsx = os.path.join(os.path.expanduser('~'), 'Desktop', 'RunFullDataSet_results.xlsx')
    df_res = pd.DataFrame(results)

    try:
        # Compute total of 'p/l liquidation trade' for summary (single-trade P/Ls)
        try:
            total_liq = float(pd.Series(df_res.get('p/l liquidation trade', [])).dropna().astype(float).sum())
        except Exception:
            total_liq = 0.0

        with pd.ExcelWriter(out_xlsx, engine='openpyxl') as writer:
            df_res.to_excel(writer, index=False, sheet_name='Results')

        # Apply coloring: green if Captured 100 points == 'Yes' else red
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        from openpyxl.utils import get_column_letter

        wb = load_workbook(out_xlsx)
        ws = wb['Results']

        green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        # Find column index for 'Captured 100 points' and color rows by Yes/No
        header = [cell.value for cell in ws[1]]
        try:
            cap_idx = header.index('Captured 100 points') + 1
        except ValueError:
            cap_idx = None

        for row in range(2, ws.max_row + 1):
            if cap_idx:
                val = ws.cell(row=row, column=cap_idx).value
                try:
                    is_yes = str(val).strip().lower() == 'yes'
                except Exception:
                    is_yes = False
                fill = green if is_yes else red
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=c).fill = fill

        # Hide the FileName column if present
        try:
            file_idx = header.index('FileName') + 1
            file_col_letter = get_column_letter(file_idx)
            ws.column_dimensions[file_col_letter].width = 0
            ws.column_dimensions[file_col_letter].hidden = True
        except ValueError:
            pass

        # If 'p/l liquidation trade' is blank for a row, fill it from 'final_P/L'
        try:
            liq_idx = header.index('p/l liquidation trade') + 1
        except ValueError:
            liq_idx = None

        try:
            final_pl_idx = header.index('final_P/L') + 1
        except ValueError:
            final_pl_idx = None

        if liq_idx and final_pl_idx:
            for row in range(2, ws.max_row + 1):
                liq_val = ws.cell(row=row, column=liq_idx).value
                if liq_val is None or str(liq_val).strip() == '':
                    ws.cell(row=row, column=liq_idx).value = ws.cell(row=row, column=final_pl_idx).value

        # Recompute total_liq after filling blanks so SUM/AVG reflect filled values
        total_liq = 0.0
        if liq_idx:
            for row in range(2, ws.max_row + 1):
                v = ws.cell(row=row, column=liq_idx).value
                try:
                    total_liq += float(v)
                except (TypeError, ValueError):
                    pass

        # Place summary label in the 'Date' column if available, else first column
        try:
            date_idx = header.index('Date') + 1
        except ValueError:
            date_idx = 1

        summary_row = ws.max_row + 1
        ws.cell(row=summary_row, column=date_idx).value = 'SUM p/l liquidation trade'
        if liq_idx:
            ws.cell(row=summary_row, column=liq_idx).value = float(total_liq)

        # Append average points per day (sum divided by number of result rows)
        try:
            n_rows = int(len(df_res)) if df_res is not None else 0
        except Exception:
            n_rows = 0
        n_rows = max(n_rows, 1)
        avg_points = float(total_liq) / n_rows
        avg_row = summary_row + 1
        ws.cell(row=avg_row, column=date_idx).value = 'AVG points per day'
        if liq_idx:
            ws.cell(row=avg_row, column=liq_idx).value = float(avg_points)

        # SUM of final_P/L column
        try:
            final_pl_sum_idx = header.index('final_P/L') + 1
        except ValueError:
            final_pl_sum_idx = None

        if final_pl_sum_idx:
            total_final_pl = 0.0
            for row in range(2, summary_row):
                v = ws.cell(row=row, column=final_pl_sum_idx).value
                try:
                    total_final_pl += float(v)
                except (TypeError, ValueError):
                    pass
            sum_final_pl_row = avg_row + 1
            ws.cell(row=sum_final_pl_row, column=date_idx).value = 'SUM final_P/L'
            ws.cell(row=sum_final_pl_row, column=final_pl_sum_idx).value = float(total_final_pl)

        # AVG of win_pct column
        try:
            win_pct_idx = header.index('win_pct') + 1
        except ValueError:
            win_pct_idx = None

        if win_pct_idx:
            win_pct_vals = []
            for row in range(2, summary_row):
                v = ws.cell(row=row, column=win_pct_idx).value
                try:
                    win_pct_vals.append(float(v))
                except (TypeError, ValueError):
                    pass
            avg_win_pct = sum(win_pct_vals) / len(win_pct_vals) if win_pct_vals else 0.0
            avg_win_pct_row = avg_row + 2
            ws.cell(row=avg_win_pct_row, column=date_idx).value = 'AVG win_pct'
            ws.cell(row=avg_win_pct_row, column=win_pct_idx).value = round(float(avg_win_pct), 2)

        # Auto-size every column to fit the widest content in that column
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                try:
                    cell_len = len(str(cell.value)) if cell.value is not None else 0
                    if cell_len > max_len:
                        max_len = cell_len
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max_len + 4

        wb.save(out_xlsx)
        print(f"Results saved to: {out_xlsx}")
    except Exception as e:
        print(f"Failed to write Excel file: {e}")
        csv_out = os.path.join(os.path.expanduser('~'), 'Desktop', 'RunFullDataSet_results.csv')
        df_res.to_csv(csv_out, index=False)
        print(f"Wrote CSV fallback to: {csv_out}")


if __name__ == '__main__':
    run_from_desktop('CBOT_MINI_YM1_ByDate_930_1000', '09:30', '10:00')
