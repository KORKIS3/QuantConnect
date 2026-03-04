#!/usr/bin/env python3
"""
Process January 2026 YM data from Desktop CSV file
Outputs results and plots to Desktop
"""

import os
import pandas as pd
from plotFigure import ChartPlotter
from RunFullDataSet import _load_csv_as_df, _compute_trade_stats

def run_january_data():
    # Read the January CSV from Desktop
    desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop', 'Jan - YM_1m_Jan26.csv')
    
    if not os.path.exists(desktop_path):
        print(f"File not found: {desktop_path}")
        return
    
    print("Loading January 2026 data...")
    # Use the robust CSV loader from RunFullDataSet
    df = _load_csv_as_df(desktop_path)
    print(f"Loaded {len(df)} rows")

    # Debug: print first few rows to understand structure
    print("\nFirst 3 rows:")
    print(df.head(3))
    print(f"\nIndex type: {type(df.index)}")
    print(f"Index timezone: {df.index.tz if hasattr(df.index, 'tz') else 'N/A'}")

    # Get unique dates in the dataset
    dates = df.index.date
    unique_dates = sorted(set(dates))
    print(f"\nFound {len(unique_dates)} unique dates")
    print(f"Date range: {unique_dates[0]} to {unique_dates[-1]}")
    
    results = []
    start_time = "09:30"
    end_time = "10:00"
    
    # Output directory for images
    image_dir = os.path.join(os.path.expanduser('~'), 'Desktop', 'TradingPics')
    os.makedirs(image_dir, exist_ok=True)
    
    # Process each date
    for target_date in unique_dates:
        date_str = target_date.strftime('%Y-%m-%d')
        print(f"\n{'='*60}")
        print(f"Processing: {date_str}")
        
        # Filter data for this specific date
        day_data = df[df.index.date == target_date].copy()

        if day_data.empty:
            print(f"No data for {date_str}")
            continue

        # Filter for time range (09:30 to 10:00)
        # First check what times are available
        try:
            available_times = [t.strftime('%H:%M:%S') for t in day_data.index.time[:5]]
            print(f"First 5 times available: {available_times}")
        except:
            pass

        try:
            # Use between_time to filter for 09:30:00 to 10:00:00
            day_data = day_data.between_time('09:30:00', '10:00:00', inclusive='both')
        except Exception as e:
            print(f"Time filtering failed: {e}")
            # Try alternative filtering method
            try:
                mask = (day_data.index.time >= pd.Timestamp('09:30:00').time()) & (day_data.index.time <= pd.Timestamp('10:00:00').time())
                day_data = day_data[mask]
            except Exception as e2:
                print(f"Alternative time filtering also failed: {e2}")

        if day_data.empty:
            print(f"No data in time range {start_time}-{end_time}")
            continue

        print(f"Data rows: {len(day_data)}")
        print(f"Time range: {day_data.index[0].strftime('%H:%M')} to {day_data.index[-1].strftime('%H:%M')}")
        
        # Create plotter and analyze
        try:
            plotter = ChartPlotter(day_data, date_str, start_time, end_time, 
                                 os.path.expanduser('~') + '/Desktop')
            plotter.create_figure()
            plotter.detect_all_signals_once()
            
            # Save figure
            img_path = os.path.join(image_dir, f"{date_str}.jpg")
            try:
                final_frame = max(0, len(day_data) - 1)
                plotter.state.current_frame = final_frame
                plotter.update_plot(final_frame)
                try:
                    plotter.fig.canvas.draw()
                    plotter.fig.canvas.flush_events()
                except Exception:
                    pass
                
                plotter.fig.savefig(img_path, dpi=150, bbox_inches='tight')
                print(f"Saved: {img_path}")
            except Exception as e:
                print(f"Failed to save image: {e}")
            
            # Close figure
            try:
                import matplotlib.pyplot as plt
                plt.close(plotter.fig)
            except Exception:
                pass
            
            # Compute stats
            stats = _compute_trade_stats(plotter)
            results.append({
                'Date': date_str,
                'final_P/L': stats['final_pl'],
                'winning_trades': stats['winning_trades'],
                'losing_trades': stats['losing_trades'],
                'win_pct': stats['win_pct'],
                'total_trades': stats['total_trades'],
                'Captured 100 points': stats.get('captured_100', 'No'),
                'p/l high': stats.get('pl_high', 0.0),
                'p/l when liquidated': stats.get('liquidation_pl', None),
                'p/l liquidation trade': stats.get('liquidation_trade_pl', None),
            })
            
        except Exception as e:
            print(f"Error processing {date_str}: {e}")
            import traceback
            traceback.print_exc()
    
    # Save results to Excel on Desktop
    if results:
        out_xlsx = os.path.join(os.path.expanduser('~'), 'Desktop', 'January_2026_Results.xlsx')
        df_res = pd.DataFrame(results)
        
        try:
            # Compute total of 'p/l liquidation trade'
            try:
                total_liq = float(pd.Series(df_res.get('p/l liquidation trade', [])).dropna().astype(float).sum())
            except Exception:
                total_liq = 0.0
            
            with pd.ExcelWriter(out_xlsx, engine='openpyxl') as writer:
                df_res.to_excel(writer, index=False, sheet_name='Results')
            
            # Apply coloring
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill
            from openpyxl.utils import get_column_letter
            
            wb = load_workbook(out_xlsx)
            ws = wb['Results']
            
            green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            
            header = [cell.value for cell in ws[1]]
            try:
                cap_idx = header.index('Captured 100 points') + 1
            except ValueError:
                cap_idx = None
            
            for row in range(2, ws.max_row + 1):
                if cap_idx:
                    val = ws.cell(row=row, column=cap_idx).value
                    try:
                        is_yes = str(val).strip().lower() == 'yes'
                    except Exception:
                        is_yes = False
                    fill = green if is_yes else red
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=c).fill = fill
            
            # Add summary rows
            try:
                liq_idx = header.index('p/l liquidation trade') + 1
            except ValueError:
                liq_idx = None
            
            try:
                date_idx = header.index('Date') + 1
            except ValueError:
                date_idx = 1
            
            summary_row = ws.max_row + 1
            ws.cell(row=summary_row, column=date_idx).value = 'SUM p/l liquidation trade'
            if liq_idx:
                ws.cell(row=summary_row, column=liq_idx).value = float(total_liq)
            
            n_rows = max(len(df_res), 1)
            avg_points = float(total_liq) / n_rows
            avg_row = summary_row + 1
            ws.cell(row=avg_row, column=date_idx).value = 'AVG points per day'
            if liq_idx:
                ws.cell(row=avg_row, column=liq_idx).value = float(avg_points)
            
            wb.save(out_xlsx)
            print(f"\n{'='*60}")
            print(f"Results saved to: {out_xlsx}")
            print(f"Images saved to: {image_dir}")
            print(f"Total liquidation P/L: {total_liq:.2f}")
            print(f"Average points per day: {avg_points:.2f}")
            print(f"{'='*60}")
        except Exception as e:
            print(f"Failed to write Excel file: {e}")
            csv_out = os.path.join(os.path.expanduser('~'), 'Desktop', 'January_2026_Results.csv')
            df_res.to_csv(csv_out, index=False)
            print(f"Wrote CSV fallback to: {csv_out}")
    else:
        print("No results to save")


if __name__ == '__main__':
    run_january_data()
