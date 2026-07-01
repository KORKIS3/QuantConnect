"""Grid comparison: every minute where ANY source has activity.
Columns: Time | Algo CSV | Backtest | IB Live | IB Fills (validation)
Rows: every minute that has a signal/fill in ANY column.
Outputs to both console and Excel.
"""
import pandas as pd
import os, pytz
from TradingAlgoFast import AlgoConfig, run_trading_algo_fast
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

_EST = pytz.timezone("US/Eastern")
_DATA_ROOT = os.path.join(os.path.expanduser("~"), "Desktop", "2YearsData", "full_day")

# --- Load live tracking CSV ---
live_path = os.path.join(os.path.expanduser("~"), "Desktop", "IB_Live", "tracking", "YM_tracking_DUO158495_2026-06-30_0930.csv")
live = pd.read_csv(live_path, index_col=0, parse_dates=True)

# --- Run backtest ---
fpath = os.path.join(_DATA_ROOT, "CBOT_MINI_YM1_2026-06-30.csv")
df = pd.read_csv(fpath, index_col=0, parse_dates=True)
df.index = pd.to_datetime(df.index, utc=True).tz_convert(_EST)
config = AlgoConfig(
    warmup_minutes=7, steep_angle_threshold=90.0, proximity_points=4.0,
    min_reversal_minutes=0, min_entry_angle=0.0, partial_tp_pts=50.0,
    wm_shield_distance=0.0, swing_anchor_threshold=10.0, cushion_points=40.0, limit_expiry_bars=5,
)
day_start = pd.Timestamp("2026-06-30 09:30", tz=_EST)
day_end = pd.Timestamp("2026-06-30 16:59", tz=_EST)
day_data = df[(df.index >= day_start) & (df.index <= day_end)]
bt = run_trading_algo_fast(day_data, "2026-06-30", "09:30", "17:00", config=config)

# --- Parse IB fills from validation report (with qty) ---
val_report = os.path.join(os.path.expanduser("~"), "Desktop", "IB_Live", "validation", "2026-06-30", "validation_report_2026-06-30.txt")
with open(val_report) as f:
    content = f.read()
in_fills = False
ib_fill_lines = []
for line in content.split("\n"):
    if line.strip().startswith("IB Fills:"):
        in_fills = True
        continue
    if in_fills:
        if line.strip() == "" or not line.startswith("  "):
            in_fills = False
            continue
        ib_fill_lines.append(line.strip())

# --- Build event map by minute for each source ---

# 1. Algo CSV signals (including partial TPs)
algo_csv_events = {}
for idx, row in live.iterrows():
    ts = pd.Timestamp(idx).floor("min")
    sig = row.get("signal", "")
    ptp_sig = row.get("partial_tp_signal", "")
    if pd.notna(sig) and sig != "":
        price = row.get("sell_price", None) if str(sig).upper() == "SELL" else row["Close"]
        if pd.isna(price):
            price = row["Close"]
        algo_csv_events[ts] = {"sig": str(sig).upper(), "price": float(price), "pl": float(row.get("session_pl", 0))}
    elif pd.notna(ptp_sig) and ptp_sig != "" and ptp_sig is not False:
        tp_price = row.get("partial_tp_price", row["Close"])
        if pd.isna(tp_price):
            tp_price = row["Close"]
        algo_csv_events[ts] = {"sig": f"TP ({ptp_sig})", "price": float(tp_price), "pl": float(row.get("session_pl", 0))}

# 2. Backtest signals (including partial TPs)
bt_events = {}
for idx, row in bt.iterrows():
    ts = idx.floor("min")
    sig = row.get("signal", "")
    ptp_sig = row.get("partial_tp_signal", "")
    if pd.notna(sig) and sig != "":
        sig_str = str(sig).upper()
        price = row.get("sell_price", None) if sig_str == "SELL" else row["Close"]
        if pd.isna(price):
            price = row["Close"]
        bt_events[ts] = {"sig": sig_str, "price": float(price), "pl": float(row.get("session_pl", 0))}
    elif pd.notna(ptp_sig) and ptp_sig != "" and ptp_sig is not False:
        tp_price = row.get("partial_tp_price", row["Close"])
        if pd.isna(tp_price):
            tp_price = row["Close"]
        bt_events[ts] = {"sig": f"TP ({ptp_sig})", "price": float(tp_price), "pl": float(row.get("session_pl", 0))}

# 3. IB Live (from ib_signal columns in tracking CSV)
ib_live_events = {}
ib_rows = live[(live["ib_signal"].notna()) & (live["ib_signal"] != "") & (live["ib_signal"] != "flat")]
for idx, row in ib_rows.iterrows():
    ts = pd.Timestamp(idx).floor("min")
    sig = row["ib_signal"]
    if sig == "BUY":
        price = row.get("ib_buy_price", row["Close"])
    else:
        price = row.get("ib_sell_price", row["Close"])
    if pd.isna(price):
        price = row["Close"]
    ib_live_events[ts] = {"sig": sig, "price": float(price), "pl": float(row.get("ib_session_pl", 0))}

# 4. IB Fills from validation report - match to ib_signal timestamps for timing
# Track position through fills to identify TPs (reduces position by 1 without flipping)
ib_fill_events = {}
ib_fill_iter = iter(ib_fill_lines)
ib_pos_tracker = 0
for ts in sorted(ib_live_events.keys()):
    try:
        fill_line = next(ib_fill_iter)
        parts = fill_line.split()
        side = "BUY" if parts[0] == "BOT" else "SELL"
        qty = int(parts[1])
        price = float(parts[3])
        
        # Determine if this fill is a TP:
        # A TP reduces position by 1 without flipping direction
        is_tp = False
        if side == "BUY" and ib_pos_tracker < 0 and qty < abs(ib_pos_tracker):
            is_tp = True  # partial close of short
        elif side == "SELL" and ib_pos_tracker > 0 and qty < ib_pos_tracker:
            is_tp = True  # partial close of long
        
        # Update position
        if side == "BUY":
            ib_pos_tracker += qty
        else:
            ib_pos_tracker -= qty
        
        ib_fill_events[ts] = {"sig": side, "qty": qty, "price": price, "raw": fill_line, "is_tp": is_tp}
    except StopIteration:
        break

# Collect remaining fills that didn't match
remaining_fills = list(ib_fill_iter)

# --- Collect all active minutes ---
all_times = set()
all_times.update(algo_csv_events.keys())
all_times.update(bt_events.keys())
all_times.update(ib_live_events.keys())
all_times.update(ib_fill_events.keys())

# --- Chart snapshots available ---
_CHART_ROOT = os.path.join(os.path.expanduser("~"), "Desktop", "IB_Live", "charts")
snapshot_times = set()
for h in [10, 11, 12, 13, 14, 15, 16]:
    snap = f"YM_2026-06-30_{h:02d}00_snapshot.jpg"
    if os.path.exists(os.path.join(_CHART_ROOT, snap)):
        snapshot_times.add(pd.Timestamp(f"2026-06-30 {h:02d}:00", tz=_EST))

# --- Print grid ---
print("=" * 130)
print("FULL GRID: Every event across all sources (2026-06-30)")
print("=" * 130)
print()

col_w = 22
hdr = (f"{'Time':<9} | "
       f"{'ALGO CSV':<{col_w}} | "
       f"{'BACKTEST':<{col_w}} | "
       f"{'IB LIVE (csv cols)':<{col_w}} | "
       f"{'IB FILLS (val rpt)':<{col_w}} | "
       f"{'Chart Snap'}")
print(hdr)
print("-" * 130)

# Track running P/L for each source
algo_pl = 0.0
bt_pl = 0.0
ib_pl = 0.0

for ts in sorted(all_times):
    time_str = ts.strftime("%H:%M")

    # Algo CSV
    if ts in algo_csv_events:
        e = algo_csv_events[ts]
        algo_str = f"{e['sig']} @ {e['price']:.0f} ({e['pl']:+.0f})"
        algo_pl = e["pl"]
    else:
        algo_str = ""

    # Backtest
    if ts in bt_events:
        e = bt_events[ts]
        bt_str = f"{e['sig']} @ {e['price']:.0f} ({e['pl']:+.0f})"
        bt_pl = e["pl"]
    else:
        bt_str = ""

    # IB Live
    if ts in ib_live_events:
        e = ib_live_events[ts]
        ib_live_str = f"{e['sig']} @ {e['price']:.0f} ({e['pl']:+.0f})"
        ib_pl = e["pl"]
    else:
        ib_live_str = ""

    # IB Fills
    if ts in ib_fill_events:
        e = ib_fill_events[ts]
        ib_fill_str = f"{e['sig']} {e['qty']} @ {e['price']:.0f}"
    else:
        ib_fill_str = ""

    # Chart snapshot
    snap_str = "YES" if ts in snapshot_times else ""

    print(f"{time_str:<9} | "
          f"{algo_str:<{col_w}} | "
          f"{bt_str:<{col_w}} | "
          f"{ib_live_str:<{col_w}} | "
          f"{ib_fill_str:<{col_w}} | "
          f"{snap_str}")

# Remaining unmatched IB fills
if remaining_fills:
    print()
    print(f"{'EXTRA':<9} | {'':^{col_w}} | {'':^{col_w}} | {'':^{col_w}} | (unmatched fills below)")
    for f in remaining_fills:
        print(f"{'  ?  ':<9} | {'':^{col_w}} | {'':^{col_w}} | {'':^{col_w}} | {f}")

# --- Hourly P/L snapshot ---
print()
print()
print("=" * 100)
print("HOURLY P/L SNAPSHOT")
print("=" * 100)
print()
print(f"{'Hour':<8} | {'Algo CSV P/L':<16} | {'Backtest P/L':<16} | {'IB Live P/L':<16} | {'Chart?':<8} | {'Algo Pos':<10} | {'IB Pos':<10}")
print("-" * 100)

hours = ["10:00", "11:00", "12:00", "13:00", "14:00", "15:00", "16:00", "16:58"]
for hour in hours:
    cutoff = pd.Timestamp(f"2026-06-30 {hour}", tz=_EST)

    # Algo CSV
    algo_slice = live[live.index <= str(cutoff)]
    a_pl = float(algo_slice["session_pl"].iloc[-1]) if not algo_slice.empty else 0
    a_pos = algo_slice["position"].iloc[-1] if not algo_slice.empty else "flat"

    # Backtest
    bt_slice = bt[bt.index <= cutoff]
    b_pl = float(bt_slice["session_pl"].iloc[-1]) if not bt_slice.empty else 0

    # IB Live
    ib_pl_val = float(algo_slice["ib_session_pl"].iloc[-1]) if (not algo_slice.empty and "ib_session_pl" in algo_slice.columns) else 0
    ib_pos = algo_slice["ib_position"].iloc[-1] if (not algo_slice.empty and "ib_position" in algo_slice.columns) else "flat"

    # Chart
    h_int = int(hour.split(":")[0])
    snap_exists = os.path.exists(os.path.join(_CHART_ROOT, f"YM_2026-06-30_{h_int:02d}00_snapshot.jpg"))
    snap_str = "YES" if snap_exists else ""

    print(f"{hour:<8} | {a_pl:>+12.0f} pts | {b_pl:>+12.0f} pts | {ib_pl_val:>+12.0f} pts | {snap_str:<8} | {a_pos:<10} | {ib_pos:<10}")

# Final summary
print()
print("=" * 100)
print("FINAL SUMMARY")
print("=" * 100)
a_final = float(live["session_pl"].iloc[-1])
b_final = float(bt["session_pl"].iloc[-1])
ib_final_csv = float(live["ib_session_pl"].iloc[-1]) if "ib_session_pl" in live.columns else 0
ib_final_fills = 266.0

print(f"  Algo CSV final:        {a_final:+.0f} pts")
print(f"  Backtest final:        {b_final:+.0f} pts")
print(f"  IB Live (csv track):   {ib_final_csv:+.0f} pts")
print(f"  IB Fills (FIFO calc):  {ib_final_fills:+.0f} pts")
print()
print(f"  Algo CSV vs Backtest:  {a_final - b_final:+.0f} pts")
print(f"  Algo CSV vs IB Fills:  {a_final - ib_final_fills:+.0f} pts")
print(f"  Backtest vs IB Fills:  {b_final - ib_final_fills:+.0f} pts")


# =============================================================================
# EXCEL OUTPUT
# =============================================================================
wb = Workbook()

# --- Sheet 1: Full Grid ---
ws1 = wb.active
ws1.title = "Full Grid"

# Styles
header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, size=11, color="FFFFFF")
buy_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
sell_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
tp_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
thick_border = Border(
    left=Side(style="medium"), right=Side(style="medium"),
    top=Side(style="medium"), bottom=Side(style="medium"),
)
section_left = Border(left=Side(style="medium"), right=Side(style="thin"),
                      top=Side(style="thin"), bottom=Side(style="thin"))
section_right = Border(left=Side(style="thin"), right=Side(style="medium"),
                       top=Side(style="thin"), bottom=Side(style="thin"))
section_mid = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))

# Headers - Row 1: Section group headers (bold, merged look)
section_headers = ["", "ALGO CSV", "", "", "", "", "", "", "BACKTEST", "", "", "", "", "", "", "IB LIVE", "", "", "", "", "", "", "IB FILLS", "", "", "", "IB LIVE SNAPSHOT", "", "", "", "", ""]
for col, h in enumerate(section_headers, 1):
    cell = ws1.cell(row=1, column=col, value=h)
    cell.font = Font(bold=True, size=12)
    cell.fill = section_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thick_border

# Headers - Row 2: Column sub-headers
headers = ["Time",
           "Signal", "Price", "Qty", "P/L", "Chg", "Total P/L",  # B-G: Algo CSV
           "",                                                      # H: blank
           "Signal", "Price", "Qty", "P/L", "Chg", "Total P/L",  # I-N: Backtest
           "",                                                      # O: blank
           "Signal", "Price", "Qty", "P/L", "Chg", "Total P/L",  # P-U: IB Live
           "",                                                      # V: blank
           "Side+Qty", "Price", "Chart",                           # W-Y: IB Fills
           "",                                                      # Z: blank
           "Trade Time", "Signal", "Fill Price", "Qty", "Chg", "Session P/L"]  # AA-AF: Snapshot
for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=2, column=col, value=h)
    if h:
        cell.font = header_font_white
        cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thick_border

# Merge section headers
from openpyxl.utils import get_column_letter
ws1.merge_cells("B1:G1")   # ALGO CSV
ws1.merge_cells("I1:N1")   # BACKTEST
ws1.merge_cells("P1:U1")   # IB LIVE
ws1.merge_cells("W1:Y1")   # IB FILLS
ws1.merge_cells("AA1:AF1") # IB LIVE SNAPSHOT

# Blank separator columns - make them narrow
for sep_col in [8, 15, 22, 26, 30]:  # H, O, V, Z
    ws1.column_dimensions[get_column_letter(sep_col)].width = 2

# Build snapshot hour lookup: which snapshot does each trade time fall into
def get_snapshot_hour(ts):
    """Return the snapshot hour this trade would appear in."""
    snap_hours_ts = [pd.Timestamp(f"2026-06-30 {h}", tz=_EST)
                     for h in ["10:00", "11:00", "12:00", "13:00", "14:00", "15:00", "16:00", "17:00"]]
    for sh in snap_hours_ts:
        if ts <= sh:
            return sh.strftime("%H:%M")
    return "16:00+"

# Build snapshot trade lookup: keyed by minute timestamp
snapshot_trades = {}
for idx, row in live.iterrows():
    ts = pd.Timestamp(idx).floor("min")
    sig = row.get("signal", "")
    ptp_sig = row.get("partial_tp_signal", "")
    if pd.notna(sig) and sig != "":
        sig_str = str(sig).upper()
        if sig_str == "SELL":
            price = row.get("sell_price", row["Close"])
        else:
            price = row["Close"]
        if pd.isna(price):
            price = row["Close"]
        spl = float(row.get("session_pl", 0))
        snapshot_trades[ts] = {
            "time": str(idx)[0:19],
            "signal": sig_str,
            "price": float(price),
            "session_pl": spl,
        }
    elif pd.notna(ptp_sig) and ptp_sig != "" and ptp_sig is not False:
        tp_price = row.get("partial_tp_price", row["Close"])
        if pd.isna(tp_price):
            tp_price = row["Close"]
        spl = float(row.get("session_pl", 0))
        snapshot_trades[ts] = {
            "time": str(idx)[0:19],
            "signal": f"TP ({ptp_sig})",
            "price": float(tp_price),
            "session_pl": spl,
        }

# Data rows
# Track running P/L for each source
algo_running_pl = 0.0
bt_running_pl = 0.0
ib_running_pl = 0.0
snap_running_pl = 0.0
algo_prev_pl = 0.0
bt_prev_pl = 0.0
ib_prev_pl = 0.0
snap_prev_pl = 0.0

row_num = 3
for ts in sorted(all_times):
    time_str = ts.strftime("%H:%M")
    time_cell = ws1.cell(row=row_num, column=1, value=time_str)
    time_cell.border = Border(left=Side(style="medium"), right=Side(style="medium"),
                              top=Side(style="thin"), bottom=Side(style="thin"))
    time_cell.font = Font(bold=True)

    # Algo CSV (cols 2-7: Signal, Price, Qty, P/L, Chg, Total P/L)
    if ts in algo_csv_events:
        e = algo_csv_events[ts]
        qty = 1 if "TP" in e["sig"] else 2
        ws1.cell(row=row_num, column=2, value=e["sig"]).border = section_left
        ws1.cell(row=row_num, column=3, value=e["price"]).border = section_mid
        ws1.cell(row=row_num, column=4, value=qty).border = section_mid
        ws1.cell(row=row_num, column=5, value=e["pl"]).border = section_mid
        chg = e["pl"] - algo_prev_pl
        ws1.cell(row=row_num, column=6, value=chg).border = section_mid
        ws1.cell(row=row_num, column=6).font = Font(bold=True, color="006100" if chg >= 0 else "9C0006")
        algo_running_pl = e["pl"]
        algo_prev_pl = e["pl"]
        ws1.cell(row=row_num, column=7, value=algo_running_pl).border = section_right
        ws1.cell(row=row_num, column=7).font = Font(bold=True)
        fill = buy_fill if e["sig"] == "BUY" else (tp_fill if "TP" in e["sig"] else sell_fill)
        for c in range(2, 8):
            ws1.cell(row=row_num, column=c).fill = fill
    else:
        for c in range(2, 8):
            ws1.cell(row=row_num, column=c).border = section_mid

    # Col 8 = blank separator

    # Backtest (cols 9-14: Signal, Price, Qty, P/L, Chg, Total P/L)
    if ts in bt_events:
        e = bt_events[ts]
        qty = 1 if "TP" in e["sig"] else 2
        ws1.cell(row=row_num, column=9, value=e["sig"]).border = section_left
        ws1.cell(row=row_num, column=10, value=e["price"]).border = section_mid
        ws1.cell(row=row_num, column=11, value=qty).border = section_mid
        ws1.cell(row=row_num, column=12, value=e["pl"]).border = section_mid
        chg = e["pl"] - bt_prev_pl
        ws1.cell(row=row_num, column=13, value=chg).border = section_mid
        ws1.cell(row=row_num, column=13).font = Font(bold=True, color="006100" if chg >= 0 else "9C0006")
        bt_running_pl = e["pl"]
        bt_prev_pl = e["pl"]
        ws1.cell(row=row_num, column=14, value=bt_running_pl).border = section_right
        ws1.cell(row=row_num, column=14).font = Font(bold=True)
        fill = buy_fill if e["sig"] == "BUY" else (tp_fill if "TP" in e["sig"] else sell_fill)
        for c in range(9, 15):
            ws1.cell(row=row_num, column=c).fill = fill
    else:
        for c in range(9, 15):
            ws1.cell(row=row_num, column=c).border = section_mid

    # Col 15 = blank separator

    # IB Live (cols 16-21: Signal, Price, Qty, P/L, Chg, Total P/L)
    if ts in ib_live_events:
        e = ib_live_events[ts]
        # Use IB Fills TP detection if same timestamp exists
        is_ib_tp = ts in ib_fill_events and ib_fill_events[ts].get("is_tp", False)
        sig_label = f"{e['sig']}_TP" if is_ib_tp else e["sig"]
        ws1.cell(row=row_num, column=16, value=sig_label).border = section_left
        ws1.cell(row=row_num, column=17, value=e["price"]).border = section_mid
        ws1.cell(row=row_num, column=18, value=1).border = section_mid
        ws1.cell(row=row_num, column=19, value=e["pl"]).border = section_mid
        chg = e["pl"] - ib_prev_pl
        ws1.cell(row=row_num, column=20, value=chg).border = section_mid
        ws1.cell(row=row_num, column=20).font = Font(bold=True, color="006100" if chg >= 0 else "9C0006")
        ib_running_pl = e["pl"]
        ib_prev_pl = e["pl"]
        ws1.cell(row=row_num, column=21, value=ib_running_pl).border = section_right
        ws1.cell(row=row_num, column=21).font = Font(bold=True)
        fill = tp_fill if is_ib_tp else (buy_fill if e["sig"] == "BUY" else sell_fill)
        for c in range(16, 22):
            ws1.cell(row=row_num, column=c).fill = fill
    else:
        for c in range(16, 22):
            ws1.cell(row=row_num, column=c).border = section_mid

    # Col 22 = blank separator

    # IB Fills (cols 23-25: Side+Qty, Price, Chart)
    if ts in ib_fill_events:
        e = ib_fill_events[ts]
        label = f"TP {e['sig']} {e['qty']}" if e.get("is_tp") else f"{e['sig']} {e['qty']}"
        ws1.cell(row=row_num, column=23, value=label).border = section_left
        ws1.cell(row=row_num, column=24, value=e["price"]).border = section_mid
        fill = tp_fill if e.get("is_tp") else (buy_fill if e["sig"] == "BUY" else sell_fill)
        ws1.cell(row=row_num, column=23).fill = fill
        ws1.cell(row=row_num, column=24).fill = fill
    else:
        ws1.cell(row=row_num, column=23).border = section_left
        ws1.cell(row=row_num, column=24).border = section_mid

    # Chart snapshot indicator (col 25)
    snap_val = "YES" if ts in snapshot_times else ""
    snap_cell = ws1.cell(row=row_num, column=25, value=snap_val)
    snap_cell.border = section_right
    if snap_val:
        snap_cell.font = Font(bold=True, color="2E75B6")

    # Col 26 = blank separator

    # IB Live Snapshot (cols 27-32: Trade Time, Signal, Fill Price, Qty, Chg, Session P/L)
    if ts in snapshot_trades:
        st = snapshot_trades[ts]
        qty = 1 if "TP" in st["signal"] else 2
        ws1.cell(row=row_num, column=27, value=st["time"]).border = section_left
        ws1.cell(row=row_num, column=28, value=st["signal"]).border = section_mid
        ws1.cell(row=row_num, column=29, value=st["price"]).border = section_mid
        ws1.cell(row=row_num, column=30, value=qty).border = section_mid
        chg = st["session_pl"] - snap_prev_pl
        ws1.cell(row=row_num, column=31, value=chg).border = section_mid
        ws1.cell(row=row_num, column=31).font = Font(bold=True, color="006100" if chg >= 0 else "9C0006")
        snap_prev_pl = st["session_pl"]
        ws1.cell(row=row_num, column=32, value=st["session_pl"]).border = section_right
        fill = buy_fill if st["signal"] == "BUY" else (tp_fill if "TP" in st["signal"] else sell_fill)
        for c in range(27, 33):
            ws1.cell(row=row_num, column=c).fill = fill
    else:
        for c in range(27, 33):
            ws1.cell(row=row_num, column=c).border = thin_border

    row_num += 1

# Final P/L totals row
row_num += 1
ws1.cell(row=row_num, column=1, value="FINAL P/L").font = Font(bold=True, size=11)
ws1.cell(row=row_num, column=1).border = thick_border
ws1.cell(row=row_num, column=7, value=algo_running_pl).border = thick_border
ws1.cell(row=row_num, column=7).font = Font(bold=True, size=11)
ws1.cell(row=row_num, column=14, value=bt_running_pl).border = thick_border
ws1.cell(row=row_num, column=14).font = Font(bold=True, size=11)
ws1.cell(row=row_num, column=21, value=ib_running_pl).border = thick_border
ws1.cell(row=row_num, column=21).font = Font(bold=True, size=11)
snap_final = float(live["session_pl"].iloc[-1])
ws1.cell(row=row_num, column=32, value=snap_final).border = thick_border
ws1.cell(row=row_num, column=32).font = Font(bold=True, size=11)

# Auto-width columns and freeze header
#        A   B   C    D   E   F    G    H  I   J    K   L   M    N    O  P   Q    R   S   T    U    V  W    X    Y   Z  AA   AB  AC  AD   AE   AF
col_widths = [8, 8, 10, 5, 9, 8, 10, 2, 8, 10, 5, 9, 8, 10, 2, 8, 10, 5, 9, 8, 10, 2, 10, 10, 7, 2, 20, 8, 10, 5, 8, 10]
from openpyxl.utils import get_column_letter
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.freeze_panes = "A3"  # freeze the two header rows

# --- Sheet 2: Hourly P/L ---
ws2 = wb.create_sheet("Hourly PL")
hourly_headers = ["Hour", "Algo CSV P/L", "Backtest P/L", "IB Live P/L", "Chart?", "Algo Position", "IB Position"]
for col, h in enumerate(hourly_headers, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

for r, hour in enumerate(hours, 2):
    cutoff = pd.Timestamp(f"2026-06-30 {hour}", tz=_EST)
    algo_slice = live[live.index <= str(cutoff)]
    a_pl = float(algo_slice["session_pl"].iloc[-1]) if not algo_slice.empty else 0
    a_pos = algo_slice["position"].iloc[-1] if not algo_slice.empty else "flat"
    bt_slice = bt[bt.index <= cutoff]
    b_pl = float(bt_slice["session_pl"].iloc[-1]) if not bt_slice.empty else 0
    ib_pl_v = float(algo_slice["ib_session_pl"].iloc[-1]) if (not algo_slice.empty and "ib_session_pl" in algo_slice.columns) else 0
    ib_pos_v = algo_slice["ib_position"].iloc[-1] if (not algo_slice.empty and "ib_position" in algo_slice.columns) else "flat"
    h_int = int(hour.split(":")[0])
    snap_exists = os.path.exists(os.path.join(_CHART_ROOT, f"YM_2026-06-30_{h_int:02d}00_snapshot.jpg"))

    ws2.cell(row=r, column=1, value=hour).border = thin_border
    ws2.cell(row=r, column=2, value=a_pl).border = thin_border
    ws2.cell(row=r, column=3, value=b_pl).border = thin_border
    ws2.cell(row=r, column=4, value=ib_pl_v).border = thin_border
    ws2.cell(row=r, column=5, value="YES" if snap_exists else "").border = thin_border
    ws2.cell(row=r, column=6, value=str(a_pos)).border = thin_border
    ws2.cell(row=r, column=7, value=str(ib_pos_v)).border = thin_border

for col in range(1, 8):
    ws2.column_dimensions[chr(64 + col)].width = 16
ws2.freeze_panes = "A2"

# Add conditional formatting — highlight rows where positions diverge
for r, hour in enumerate(hours, 2):
    algo_pos_val = ws2.cell(row=r, column=6).value
    ib_pos_val = ws2.cell(row=r, column=7).value
    if algo_pos_val != ib_pos_val and algo_pos_val and ib_pos_val:
        warn_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        for c in range(1, 8):
            ws2.cell(row=r, column=c).fill = warn_fill

# --- Sheet 3: Summary ---
ws3 = wb.create_sheet("Summary")
a_final = float(live["session_pl"].iloc[-1])
b_final = float(bt["session_pl"].iloc[-1])
ib_csv_final = float(live["ib_session_pl"].iloc[-1]) if "ib_session_pl" in live.columns else 0
ib_final = 266.0
summary_data = [
    ["Source", "Signals", "Final P/L (pts)", "Final P/L (USD)"],
    ["Algo CSV", 17, a_final, a_final * 5],
    ["Backtest", 17, b_final, b_final * 5],
    ["IB Live (csv track)", 23, ib_csv_final, ib_csv_final * 5],
    ["IB Fills (FIFO)", 22, ib_final, ib_final * 5],
    [],
    ["Comparison", "", "Difference (pts)", ""],
    ["Algo CSV vs Backtest", "", a_final - b_final, ""],
    ["Algo CSV vs IB Fills", "", a_final - ib_final, ""],
    ["Backtest vs IB Fills", "", b_final - ib_final, ""],
]
for r, row_data in enumerate(summary_data, 1):
    for c, val in enumerate(row_data, 1):
        cell = ws3.cell(row=r, column=c, value=val)
        cell.border = thin_border
        if r == 1 or r == 7:
            cell.font = header_font_white
            cell.fill = header_fill

for col in range(1, 5):
    ws3.column_dimensions[chr(64 + col)].width = 22

# --- Sheet 4: Chart Snapshot Details ---
ws4 = wb.create_sheet("Chart Snapshots")

# Header
snap_headers = ["Hour", "Image?", "Algo Position", "Algo P/L", "IB Position", "IB P/L",
                "Signals", "Buys", "Sells", "Close Price"]
for col, h in enumerate(snap_headers, 1):
    cell = ws4.cell(row=1, column=col, value=h)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thick_border

snapshot_hours_full = ["09:30", "10:00", "11:00", "12:00", "13:00", "14:00", "15:00", "16:00"]
row_num_snap = 2
for hour in snapshot_hours_full:
    cutoff = pd.Timestamp(f"2026-06-30 {hour}", tz=_EST)
    sliced = live[live.index <= str(cutoff)]
    if sliced.empty:
        continue

    signals_slice = sliced[(sliced["signal"].notna()) & (sliced["signal"] != "")]
    n_buys = len(signals_slice[signals_slice["signal"].str.upper() == "BUY"])
    n_sells = len(signals_slice[signals_slice["signal"].str.upper() == "SELL"])
    last_row = sliced.iloc[-1]
    algo_pos = last_row.get("position", "flat")
    algo_pl_val = float(last_row.get("session_pl", 0))
    ib_pos_val = last_row.get("ib_position", "flat") if "ib_position" in sliced.columns else "flat"
    ib_pl_snap = float(last_row.get("ib_session_pl", 0)) if "ib_session_pl" in sliced.columns else 0
    close_p = float(last_row["Close"])

    h_int = int(hour.split(":")[0])
    m_int = int(hour.split(":")[1])
    if hour == "09:30":
        img_name = "YM_2026-06-30_0930.jpg"
    else:
        img_name = f"YM_2026-06-30_{h_int:02d}{m_int:02d}_snapshot.jpg"
    img_exists = os.path.exists(os.path.join(_CHART_ROOT, img_name))

    ws4.cell(row=row_num_snap, column=1, value=hour).border = thin_border
    ws4.cell(row=row_num_snap, column=2, value="YES" if img_exists else "NO").border = thin_border
    ws4.cell(row=row_num_snap, column=3, value=str(algo_pos)).border = thin_border
    ws4.cell(row=row_num_snap, column=4, value=algo_pl_val).border = thin_border
    ws4.cell(row=row_num_snap, column=5, value=str(ib_pos_val)).border = thin_border
    ws4.cell(row=row_num_snap, column=6, value=ib_pl_snap).border = thin_border
    ws4.cell(row=row_num_snap, column=7, value=len(signals_slice)).border = thin_border
    ws4.cell(row=row_num_snap, column=8, value=n_buys).border = thin_border
    ws4.cell(row=row_num_snap, column=9, value=n_sells).border = thin_border
    ws4.cell(row=row_num_snap, column=10, value=close_p).border = thin_border

    # Highlight position mismatch
    if str(algo_pos) != str(ib_pos_val):
        warn_fill_snap = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        for c in range(1, 11):
            ws4.cell(row=row_num_snap, column=c).fill = warn_fill_snap

    row_num_snap += 1

# Totals row for the snapshot summary table
ws4.cell(row=row_num_snap, column=1, value="FINAL").border = thick_border
ws4.cell(row=row_num_snap, column=1).font = Font(bold=True)
ws4.cell(row=row_num_snap, column=2, value="").border = thick_border
ws4.cell(row=row_num_snap, column=3, value=str(live["position"].iloc[-1])).border = thick_border
ws4.cell(row=row_num_snap, column=4, value=float(live["session_pl"].iloc[-1])).border = thick_border
ib_pos_final = live["ib_position"].iloc[-1] if "ib_position" in live.columns else "flat"
ib_pl_final = float(live["ib_session_pl"].iloc[-1]) if "ib_session_pl" in live.columns else 0
ws4.cell(row=row_num_snap, column=5, value=str(ib_pos_final)).border = thick_border
ws4.cell(row=row_num_snap, column=6, value=ib_pl_final).border = thick_border
total_sigs = len(live[(live["signal"].notna()) & (live["signal"] != "")])
ws4.cell(row=row_num_snap, column=7, value=total_sigs).border = thick_border
total_buys = len(live[(live["signal"].notna()) & (live["signal"].str.upper() == "BUY")])
total_sells = len(live[(live["signal"].notna()) & (live["signal"].str.upper() == "SELL")])
ws4.cell(row=row_num_snap, column=8, value=total_buys).border = thick_border
ws4.cell(row=row_num_snap, column=9, value=total_sells).border = thick_border
ws4.cell(row=row_num_snap, column=10, value=float(live["Close"].iloc[-1])).border = thick_border
for c in range(1, 11):
    ws4.cell(row=row_num_snap, column=c).font = Font(bold=True)
row_num_snap += 1

# Add trade detail rows below the summary
row_num_snap += 1
ws4.cell(row=row_num_snap, column=1, value="TRADES VISIBLE ON EACH CHART").font = Font(bold=True, size=12)
row_num_snap += 1

trade_headers = ["Snapshot Hour", "Trade Time", "Signal", "Fill Price", "Session P/L"]
for col, h in enumerate(trade_headers, 1):
    cell = ws4.cell(row=row_num_snap, column=col, value=h)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.border = thick_border
row_num_snap += 1

for hour in snapshot_hours_full:
    cutoff = pd.Timestamp(f"2026-06-30 {hour}", tz=_EST)
    sliced = live[live.index <= str(cutoff)]
    if sliced.empty:
        continue
    signals_slice = sliced[(sliced["signal"].notna()) & (sliced["signal"] != "")]
    if signals_slice.empty:
        continue

    # Only show NEW trades since previous snapshot
    prev_hour_idx = snapshot_hours_full.index(hour) - 1
    if prev_hour_idx >= 0:
        prev_cutoff = pd.Timestamp(f"2026-06-30 {snapshot_hours_full[prev_hour_idx]}", tz=_EST)
        new_signals = signals_slice[signals_slice.index > str(prev_cutoff)]
    else:
        new_signals = signals_slice

    if new_signals.empty:
        continue

    # Bold separator row
    sep_cell = ws4.cell(row=row_num_snap, column=1,
                        value=f"--- {hour} ({len(signals_slice)} total, {len(new_signals)} new since last) ---")
    sep_cell.font = Font(bold=True, size=10)
    ws4.merge_cells(start_row=row_num_snap, start_column=1, end_row=row_num_snap, end_column=5)
    row_num_snap += 1

    for idx, row in new_signals.iterrows():
        sig = str(row["signal"]).upper()
        if sig == "SELL":
            price = row.get("sell_price", row["Close"])
        else:
            price = row["Close"]
        if pd.isna(price):
            price = row["Close"]
        spl = float(row.get("session_pl", 0))

        ws4.cell(row=row_num_snap, column=1, value=hour).border = thin_border
        ws4.cell(row=row_num_snap, column=2, value=str(idx)[0:19]).border = thin_border
        ws4.cell(row=row_num_snap, column=3, value=sig).border = thin_border
        ws4.cell(row=row_num_snap, column=4, value=float(price)).border = thin_border
        ws4.cell(row=row_num_snap, column=5, value=spl).border = thin_border

        fill_color = buy_fill if sig == "BUY" else sell_fill
        for c in range(1, 6):
            ws4.cell(row=row_num_snap, column=c).fill = fill_color
        row_num_snap += 1

    # Subtotal row for this snapshot hour
    last_pl = float(signals_slice.iloc[-1].get("session_pl", 0))
    ws4.cell(row=row_num_snap, column=1, value="").border = thin_border
    ws4.cell(row=row_num_snap, column=2, value="").border = thin_border
    ws4.cell(row=row_num_snap, column=3, value="").border = thin_border
    ws4.cell(row=row_num_snap, column=4, value=f"P/L @ {hour}:").border = thin_border
    ws4.cell(row=row_num_snap, column=4).font = Font(bold=True)
    ws4.cell(row=row_num_snap, column=5, value=last_pl).border = thin_border
    ws4.cell(row=row_num_snap, column=5).font = Font(bold=True)
    row_num_snap += 1

    row_num_snap += 1

# P/L totals row at the bottom
row_num_snap += 1
ws4.cell(row=row_num_snap, column=1, value="").border = thin_border
ws4.cell(row=row_num_snap, column=2, value="").border = thin_border
ws4.cell(row=row_num_snap, column=3, value="").border = thin_border
ws4.cell(row=row_num_snap, column=4, value="TOTAL DAY P/L:").border = thick_border
ws4.cell(row=row_num_snap, column=4).font = Font(bold=True, size=12)
ws4.cell(row=row_num_snap, column=5, value=float(live["session_pl"].iloc[-1])).border = thick_border
ws4.cell(row=row_num_snap, column=5).font = Font(bold=True, size=12)
row_num_snap += 1

for col in range(1, 11):
    ws4.column_dimensions[chr(64 + col)].width = 18
ws4.freeze_panes = "A2"

# Save
output_path = os.path.join(os.path.expanduser("~"), "Desktop", "IB_Live", "validation", "2026-06-30", "comparison_grid_2026-06-30.xlsx")
os.makedirs(os.path.dirname(output_path), exist_ok=True)
wb.save(output_path)
print(f"\n\nExcel saved: {output_path}")
