"""
Process YM_1m_Jan26_parsed.csv data with RunFullDataSet logic
Generates charts for each day in the 10:00-11:00 AM window
Calculates total P/L and average P/L per day
"""
import pandas as pd
import pytz
import os
import datetime
from plotFigure import ChartPlotter
from RunFullDataSet import _compute_trade_stats

# Load the parsed CSV
csv_file = os.path.join(os.path.expanduser('~'), 'Desktop', 'YM_1m_Jan26_parsed.csv')
print(f"Loading {csv_file}...")

df = pd.read_csv(csv_file, parse_dates=['Timestamp'])
df = df.set_index('Timestamp')

# Ensure timezone is set
est = pytz.timezone('US/Eastern')
if df.index.tz is None:
    df.index = df.index.tz_localize(est, ambiguous='NaT', nonexistent='shift_forward')
else:
    df.index = df.index.tz_convert(est)

print(f"Loaded {len(df)} rows")
print(f"Date range: {df.index[0]} to {df.index[-1]}")

# Group by date
df['date'] = df.index.date
dates = sorted(df['date'].unique())
print(f"Found {len(dates)} trading days\n")

# Process each day with 10:00-11:00 window
output_dir = os.path.join(os.path.expanduser('~'), 'Desktop', 'Trading', 'Temp')
os.makedirs(output_dir, exist_ok=True)

results = []
start_time = "10:00"
end_time = "11:00"

for date in dates:
    date_str = date.strftime('%Y-%m-%d')
    print(f"\n{'='*60}")
    print(f"Processing: {date_str}")
    
    # Filter data for this date
    df_day = df[df['date'] == date].copy()
    df_day = df_day.drop('date', axis=1)
    
    if df_day.empty:
        print(f"  Skipped: no data")
        continue
    
    # Filter to regular trading hours (9:30 AM - 4:00 PM)
    morning_cutoff = datetime.time(9, 30)
    afternoon_cutoff = datetime.time(16, 0)
    df_day = df_day[(df_day.index.time >= morning_cutoff) & (df_day.index.time <= afternoon_cutoff)]
    
    if df_day.empty:
        print(f"  Skipped: no regular trading hours data")
        continue
    
    # Filter to 10:00-11:00 window for chart
    chart_start = datetime.time(10, 0)
    chart_end = datetime.time(11, 0)
    df_window = df_day[(df_day.index.time >= chart_start) & (df_day.index.time <= chart_end)]
    
    if len(df_window) < 10:
        print(f"  Skipped: only {len(df_window)} minutes of data")
        continue
    
    print(f"  Window: {len(df_window)} minutes from {df_window.index[0].strftime('%H:%M')} to {df_window.index[-1].strftime('%H:%M')}")
    
    # Create chart
    try:
        plotter = ChartPlotter(df_window, date_str, start_time, end_time, output_dir)
        plotter.create_figure()
        plotter.detect_all_signals_once()
        
        # Save image
        image_dir = os.path.join(os.path.expanduser('~'), 'Desktop', 'TradingPics_Jan26')
        os.makedirs(image_dir, exist_ok=True)
        img_path = os.path.join(image_dir, f"{date_str}.jpg")
        
        if hasattr(plotter, 'fig'):
            final_frame = max(0, len(df_window) - 1)
            plotter.state.current_frame = final_frame
            plotter.update_plot(final_frame)
            
            try:
                plotter.fig.canvas.draw()
                plotter.fig.canvas.flush_events()
            except:
                pass
            
            plotter.fig.savefig(img_path, dpi=150, bbox_inches='tight')
            print(f"  ✓ Saved: {img_path}")
            
            import matplotlib.pyplot as plt
            plt.close(plotter.fig)
        
        # Collect stats
        stats = _compute_trade_stats(plotter)

        results.append({
            'Date': date_str,
            'Minutes': len(df_window),
            'Buy Signals': len(plotter.state.detected_buy_signals),
            'Sell Signals': len(plotter.state.detected_sell_signals),
            'Total Trades': stats['total_trades'],
            'Winning Trades': stats['winning_trades'],
            'Losing Trades': stats['losing_trades'],
            'Win %': stats['win_pct'],
            'Final P/L': stats['final_pl'],
            'Captured 100 pts': stats['captured_100'],
            'P/L High': stats['pl_high'],
            'P/L Liquidation': stats.get('liquidation_trade_pl', None)
        })
        
    except Exception as e:
        print(f"  ✗ Error: {e}")
        import traceback
        traceback.print_exc()

print(f"\n{'='*60}")
print(f"✓ Complete! Processed {len(results)} days")
print(f"Charts saved to Desktop/TradingPics_Jan26/")

# Calculate summary statistics
summary_df = pd.DataFrame(results)

# Calculate totals
total_pl = summary_df['Final P/L'].sum()
total_liquidation_pl = summary_df['P/L Liquidation'].dropna().sum()
avg_pl_per_day = total_pl / len(results) if len(results) > 0 else 0
avg_liquidation_per_day = total_liquidation_pl / len(results) if len(results) > 0 else 0

print(f"\n{'='*60}")
print(f"📊 JANUARY 2026 SUMMARY")
print(f"{'='*60}")
print(f"Trading Days: {len(results)}")
print(f"Total P/L: ${total_pl:,.2f}")
print(f"Total Liquidation P/L: ${total_liquidation_pl:,.2f}")
print(f"Average P/L per Day: ${avg_pl_per_day:,.2f}")
print(f"Average Liquidation P/L per Day: ${avg_liquidation_per_day:,.2f}")
print(f"{'='*60}")

# Save detailed summary to Excel with formatting
summary_file = os.path.join(os.path.expanduser('~'), 'Desktop', 'Jan26_Summary.xlsx')

try:
    with pd.ExcelWriter(summary_file, engine='openpyxl') as writer:
        summary_df.to_excel(writer, index=False, sheet_name='Daily Results')

    # Apply formatting
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = load_workbook(summary_file)
    ws = wb['Daily Results']

    # Color coding for Captured 100 pts column
    green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    header = [cell.value for cell in ws[1]]
    try:
        cap_idx = header.index('Captured 100 pts') + 1
    except ValueError:
        cap_idx = None

    # Color rows based on Captured 100 pts
    for row in range(2, ws.max_row + 1):
        if cap_idx:
            val = ws.cell(row=row, column=cap_idx).value
            is_yes = str(val).strip().lower() == 'yes'
            fill = green if is_yes else red
            for c in range(1, ws.max_column + 1):
                ws.cell(row=row, column=c).fill = fill

    # Add summary rows
    summary_start_row = ws.max_row + 2

    # Headers for summary
    ws.cell(row=summary_start_row, column=1).value = "SUMMARY"
    ws.cell(row=summary_start_row, column=1).font = Font(bold=True, size=14)

    # Total P/L
    ws.cell(row=summary_start_row + 1, column=1).value = "Total P/L"
    ws.cell(row=summary_start_row + 1, column=2).value = float(total_pl)
    ws.cell(row=summary_start_row + 1, column=1).font = Font(bold=True)

    # Total Liquidation P/L
    ws.cell(row=summary_start_row + 2, column=1).value = "Total Liquidation P/L"
    ws.cell(row=summary_start_row + 2, column=2).value = float(total_liquidation_pl)
    ws.cell(row=summary_start_row + 2, column=1).font = Font(bold=True)

    # Average P/L per Day
    ws.cell(row=summary_start_row + 3, column=1).value = "Average P/L per Day"
    ws.cell(row=summary_start_row + 3, column=2).value = float(avg_pl_per_day)
    ws.cell(row=summary_start_row + 3, column=1).font = Font(bold=True)

    # Average Liquidation P/L per Day
    ws.cell(row=summary_start_row + 4, column=1).value = "Average Liquidation P/L per Day"
    ws.cell(row=summary_start_row + 4, column=2).value = float(avg_liquidation_per_day)
    ws.cell(row=summary_start_row + 4, column=1).font = Font(bold=True)

    # Number of days
    ws.cell(row=summary_start_row + 5, column=1).value = "Trading Days"
    ws.cell(row=summary_start_row + 5, column=2).value = len(results)
    ws.cell(row=summary_start_row + 5, column=1).font = Font(bold=True)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(summary_file)
    print(f"\n✓ Summary saved to: {summary_file}")

except Exception as e:
    print(f"\n✗ Failed to create formatted Excel: {e}")
    # Fallback to simple CSV
    summary_file_csv = os.path.join(os.path.expanduser('~'), 'Desktop', 'Jan26_Summary.csv')
    summary_df.to_csv(summary_file_csv, index=False)
    print(f"✓ CSV fallback saved to: {summary_file_csv}")

print(f"{'='*60}")
