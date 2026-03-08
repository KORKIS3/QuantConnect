"""RunAllDays

Entry script to run the YM trading workflow for all days in a folder.

This module scans the standard desktop folder of per-day YM CSV files
and calls `ReOrgMain.run_single_day` for each date in order.

In addition to per-day images/CSVs it also produces a summary Excel
workbook (and CSV fallback) similar to the original
`RunFullDataSet_results.xlsx`, written as `RunAllDaysOutput.xlsx` on
the desktop.
"""

import glob
import os
import re

import pandas as pd

from ReOrgMain import run_single_day


def _extract_date_from_filename(fname: str) -> str:
    """Best-effort extraction of YYYY-MM-DD from the CSV filename."""

    m = re.search(r"(\d{4}-\d{2}-\d{2})", fname)
    if m:
        return m.group(1)

    m2 = re.search(r"(\d{8})", fname)
    if m2:
        s = m2.group(1)
        return f"{s[0:4]}-{s[4:6]}-{s[6:8]}"

    # Fallback: leave as empty string; caller may skip.
    return ""


def _compute_trade_stats_from_df(df: pd.DataFrame) -> dict:
    """Compute trade statistics from the tracking DataFrame.

    This mirrors the behaviour of `_compute_trade_stats` in
    `RunFullDataSet.py`, but derives buy/sell events from the
    `signal`/`buy_price`/`sell_price` columns produced by
    `TradingAlgo.run_trading_algo`.
    """

    import pytz

    est = pytz.timezone("US/Eastern")

    # Build events list from per-minute signals
    events = []
    for ts, row in df.iterrows():
        sig = str(row.get("signal", "")).strip().upper()
        if sig == "BUY":
            price = row.get("buy_price")
            try:
                price_f = float(price) if price is not None and price == price else float(row["Close"])
            except Exception:
                price_f = float(row["Close"])
            events.append((ts, "buy", price_f))
        elif sig == "SELL":
            price = row.get("sell_price")
            try:
                price_f = float(price) if price is not None and price == price else float(row["Close"])
            except Exception:
                price_f = float(row["Close"])
            events.append((ts, "sell", price_f))

    events.sort(key=lambda x: x[0])

    # Normalize event times to df.index tz
    events_by_time = {}
    for t, typ, price in events:
        try:
            evt = pd.to_datetime(t)
            if evt.tzinfo is None:
                evt = evt.tz_localize(est)
            else:
                evt = evt.tz_convert(est)
        except Exception:
            evt = pd.to_datetime(t)
        events_by_time.setdefault(evt, []).append((typ, price))

    trades = []
    position = "flat"
    entry_price = None
    entry_time = None

    cumulative_realized_pl = 0.0
    max_total_pl = 0.0
    captured100 = False
    liquidation_p_l = None
    liquidation_trade_pl = None

    # iterate through each timestamp in the data and process events at that time
    for ts in df.index:
        evs = events_by_time.get(ts, [])
        for typ, price in evs:
            if typ == "buy":
                if position == "flat":
                    position = "long"
                    entry_price = price
                    entry_time = ts
                elif position == "short":
                    pl = entry_price - price
                    cumulative_realized_pl += pl
                    trades.append({"entry_time": entry_time, "exit_time": ts, "pl": pl})
                    if pl >= 100 and liquidation_p_l is None:
                        liquidation_p_l = cumulative_realized_pl
                        liquidation_trade_pl = float(pl)
                    position = "long"
                    entry_price = price
                    entry_time = ts
            else:  # sell
                if position == "flat":
                    position = "short"
                    entry_price = price
                    entry_time = ts
                elif position == "long":
                    pl = price - entry_price
                    cumulative_realized_pl += pl
                    trades.append({"entry_time": entry_time, "exit_time": ts, "pl": pl})
                    if pl >= 100 and liquidation_p_l is None:
                        liquidation_p_l = cumulative_realized_pl
                        liquidation_trade_pl = float(pl)
                    position = "short"
                    entry_price = price
                    entry_time = ts

        # compute current total P/L (realized + unrealized)
        try:
            current_close = float(df["Close"].loc[ts])
        except Exception:
            current_close = float(df["Close"].iloc[-1])

        unrealized = 0.0
        if position == "long" and entry_price is not None:
            unrealized = current_close - entry_price
        elif position == "short" and entry_price is not None:
            unrealized = entry_price - current_close

        total_pl = cumulative_realized_pl + unrealized
        if total_pl > max_total_pl:
            max_total_pl = total_pl
        if (not captured100) and total_pl >= 100.0:
            captured100 = True

    # at end of data, if still in a position, close at last price
    if position != "flat" and entry_price is not None:
        last_price = float(df["Close"].iloc[-1])
        if position == "long":
            pl = last_price - entry_price
        else:
            pl = entry_price - last_price
        cumulative_realized_pl += pl
        trades.append({"entry_time": entry_time, "exit_time": df.index[-1], "pl": pl})
        if pl >= 100 and liquidation_p_l is None:
            liquidation_p_l = cumulative_realized_pl
            liquidation_trade_pl = float(pl)

    total_pl = cumulative_realized_pl
    wins = sum(1 for t in trades if t["pl"] > 0)
    losses = sum(1 for t in trades if t["pl"] <= 0)
    total_trades = len(trades)
    win_pct = (wins / total_trades * 100) if total_trades > 0 else 0.0

    captured_by_trade = any(t["pl"] >= 100 for t in trades)
    captured_flag = captured100 or captured_by_trade

    return {
        "final_pl": total_pl,
        "winning_trades": wins,
        "losing_trades": losses,
        "win_pct": win_pct,
        "total_trades": total_trades,
        "captured_100": "Yes" if captured_flag else "No",
        "pl_high": float(max_total_pl),
        "liquidation_pl": float(liquidation_p_l) if liquidation_p_l is not None else None,
        "liquidation_trade_pl": float(liquidation_trade_pl) if liquidation_trade_pl is not None else None,
    }


def _write_run_all_days_results(results: list, out_dir: str | None = None) -> None:
    """Write aggregated results to summary files.

    This closely follows the formatting and summary logic used in
    `RunFullDataSet.py`, but writes to `RunAllDays_results.xlsx`.

    If `out_dir` is provided the files are written there; otherwise they
    fall back to the Desktop.

    For convenience, a plain CSV `RunAllDays.csv` is also written.
    """

    df_res = pd.DataFrame(results)

    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    if out_dir is None:
        out_dir = desktop
    os.makedirs(out_dir, exist_ok=True)
    out_xlsx = os.path.join(out_dir, "RunAllDays_results.xlsx")

    try:
        # Compute total of 'p/l liquidation trade' for summary
        try:
            total_liq = float(pd.Series(df_res.get("p/l liquidation trade", [])).dropna().astype(float).sum())
        except Exception:
            total_liq = 0.0

        with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
            df_res.to_excel(writer, index=False, sheet_name="Results")

        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        from openpyxl.utils import get_column_letter

        wb = load_workbook(out_xlsx)
        ws = wb["Results"]

        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        header = [cell.value for cell in ws[1]]
        try:
            cap_idx = header.index("Captured 100 points") + 1
        except ValueError:
            cap_idx = None

        for row in range(2, ws.max_row + 1):
            if cap_idx:
                val = ws.cell(row=row, column=cap_idx).value
                try:
                    is_yes = str(val).strip().lower() == "yes"
                except Exception:
                    is_yes = False
                fill = green if is_yes else red
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=c).fill = fill

        # Hide the FileName column if present
        try:
            file_idx = header.index("FileName") + 1
            file_col_letter = get_column_letter(file_idx)
            ws.column_dimensions[file_col_letter].width = 0
            ws.column_dimensions[file_col_letter].hidden = True
        except ValueError:
            pass

        # If 'p/l liquidation trade' is blank, fill from 'final_P/L'
        try:
            liq_idx = header.index("p/l liquidation trade") + 1
        except ValueError:
            liq_idx = None

        try:
            final_pl_idx = header.index("final_P/L") + 1
        except ValueError:
            final_pl_idx = None

        if liq_idx and final_pl_idx:
            for row in range(2, ws.max_row + 1):
                liq_val = ws.cell(row=row, column=liq_idx).value
                if liq_val is None or str(liq_val).strip() == "":
                    ws.cell(row=row, column=liq_idx).value = ws.cell(row=row, column=final_pl_idx).value

        # Recompute total_liq after filling blanks
        total_liq = 0.0
        if liq_idx:
            for row in range(2, ws.max_row + 1):
                v = ws.cell(row=row, column=liq_idx).value
                try:
                    total_liq += float(v)
                except (TypeError, ValueError):
                    pass

        # Place summary label in the 'Date' column if available
        try:
            date_idx = header.index("Date") + 1
        except ValueError:
            date_idx = 1

        summary_row = ws.max_row + 1
        ws.cell(row=summary_row, column=date_idx).value = "SUM p/l liquidation trade"
        if liq_idx:
            ws.cell(row=summary_row, column=liq_idx).value = float(total_liq)

        # Average points per day
        try:
            n_rows = int(len(df_res)) if df_res is not None else 0
        except Exception:
            n_rows = 0
        n_rows = max(n_rows, 1)
        avg_points = float(total_liq) / n_rows
        avg_row = summary_row + 1
        ws.cell(row=avg_row, column=date_idx).value = "AVG points per day"
        if liq_idx:
            ws.cell(row=avg_row, column=liq_idx).value = float(avg_points)

        # SUM of final_P/L column
        try:
            final_pl_sum_idx = header.index("final_P/L") + 1
        except ValueError:
            final_pl_sum_idx = None

        if final_pl_sum_idx:
            total_final_pl = 0.0
            for row in range(2, summary_row):
                v = ws.cell(row=row, column=final_pl_sum_idx).value
                try:
                    total_final_pl += float(v)
                except (TypeError, ValueError):
                    pass
            sum_final_pl_row = avg_row + 1
            ws.cell(row=sum_final_pl_row, column=date_idx).value = "SUM final_P/L"
            ws.cell(row=sum_final_pl_row, column=final_pl_sum_idx).value = float(total_final_pl)

        # AVG of win_pct column
        try:
            win_pct_idx = header.index("win_pct") + 1
        except ValueError:
            win_pct_idx = None

        if win_pct_idx:
            win_pct_vals = []
            for row in range(2, summary_row):
                v = ws.cell(row=row, column=win_pct_idx).value
                try:
                    win_pct_vals.append(float(v))
                except (TypeError, ValueError):
                    pass
            avg_win_pct = sum(win_pct_vals) / len(win_pct_vals) if win_pct_vals else 0.0
            avg_win_pct_row = avg_row + 2
            ws.cell(row=avg_win_pct_row, column=date_idx).value = "AVG win_pct"
            ws.cell(row=avg_win_pct_row, column=win_pct_idx).value = round(float(avg_win_pct), 2)

        # Auto-size columns
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                try:
                    cell_len = len(str(cell.value)) if cell.value is not None else 0
                    if cell_len > max_len:
                        max_len = cell_len
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max_len + 4

        wb.save(out_xlsx)
        print(f"Results saved to: {out_xlsx}")

        # Always emit a simple CSV summary as well.
        csv_simple = os.path.join(out_dir, "RunAllDays.csv")
        df_res.to_csv(csv_simple, index=False)
        print(f"Simple CSV summary written to: {csv_simple}")
    except Exception as e:
        print(f"Failed to write Excel file: {e}")
        csv_out = os.path.join(out_dir, "RunAllDays_results.csv")
        df_res.to_csv(csv_out, index=False)
        print(f"Wrote CSV fallback to: {csv_out}")

        # Still ensure `RunAllDays.csv` exists for the user.
        csv_simple = os.path.join(out_dir, "RunAllDays.csv")
        try:
            df_res.to_csv(csv_simple, index=False)
            print(f"Simple CSV summary written to: {csv_simple}")
        except Exception:
            pass


def run_all_days(
    start_time: str = "09:30",
    end_time: str = "10:00",
    csv_root: str | None = None,
) -> None:
    """Loop over all CSV days in the folder and run the workflow."""

    if csv_root is None:
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        csv_root = os.path.join(desktop, "CBOT_MINI_YM1_ByDate_930_1000")

    # Where to store batch chart images and tracking CSVs on the desktop.
    desktop_root = os.path.join(os.path.expanduser("~"), "Desktop")
    image_root = os.path.join(desktop_root, "TradingPics")
    tracking_root = os.path.join(desktop_root, "TradingTracking")

    pattern = os.path.join(csv_root, "*.csv")
    files = sorted(glob.glob(pattern))
    if not files:
        print(f"RunAllDays: no CSV files found in {csv_root}")
        return

    results: list[dict] = []

    for fp in files:
        fname = os.path.basename(fp)
        target_date = _extract_date_from_filename(fname)
        if not target_date:
            print(f"RunAllDays: skipping {fname} (no parsable date)")
            continue

        print("\n" + "=" * 60)
        print(f"RunAllDays: processing {fname} (date={target_date})")
        print("=" * 60)

        try:
            # Headless run: no interactive graph. We only want a final
            # chart image per day and a per-minute tracking CSV.
            algo_df = run_single_day(
                target_date,
                start_time,
                end_time,
                csv_root=csv_root,
                show_plot=False,
                image_root=image_root,
                tracking_root=tracking_root,
            )

            # Collect per-day stats for the summary workbook/CSV.
            stats = _compute_trade_stats_from_df(algo_df)
            results.append({
                "FileName": fname,
                "Date": target_date,
                "final_P/L": stats["final_pl"],
                "winning_trades": stats["winning_trades"],
                "losing_trades": stats["losing_trades"],
                "win_pct": stats["win_pct"],
                "Captured 100 points": stats.get("captured_100", "No"),
                "p/l high": stats.get("pl_high", 0.0),
                "p/l liquidation trade": stats.get("liquidation_trade_pl", None),
            })
        except Exception as exc:  # keep running other days
            print(f"RunAllDays: error processing {fname}: {exc}")

    if results:
        _write_run_all_days_results(results)


if __name__ == "__main__":

    time_periods = [
        ("09:30", "10:00", "CBOT_MINI_YM1_ByDate_930_1000"),
        ("10:00", "10:30", "CBOT_MINI_YM1_ByDate_1000_1030"),
        ("10:30", "11:00", "CBOT_MINI_YM1_ByDate_1030_1100"),
        ("11:00", "11:30", "CBOT_MINI_YM1_ByDate_1100_1130"),
    ]

    desktop_root = os.path.join(os.path.expanduser("~"), "Desktop")
    data_root = os.path.join(desktop_root, "data")

    for start_time, end_time, folder_name in time_periods:
        period_label = f"{start_time.replace(':','')}-{end_time.replace(':','')}"
        image_root = os.path.join(desktop_root, "TradingPics", period_label)
        tracking_root = os.path.join(desktop_root, "TradingTracking", period_label)
        os.makedirs(image_root, exist_ok=True)
        os.makedirs(tracking_root, exist_ok=True)

        csv_root = os.path.join(data_root, folder_name)
        results = []
        pattern = os.path.join(csv_root, "*.csv")
        files = sorted(glob.glob(pattern))
        if not files:
            print(f"RunAllDays: no CSV files found in {csv_root}")
            continue

        for fp in files:
            fname = os.path.basename(fp)
            target_date = _extract_date_from_filename(fname)
            if not target_date:
                print(f"RunAllDays: skipping {fname} (no parsable date)")
                continue

            print("\n" + "=" * 60)
            print(f"RunAllDays: processing {fname} (date={target_date}) [{start_time}-{end_time}]")
            print("=" * 60)

            try:
                algo_df = run_single_day(
                    target_date,
                    start_time,
                    end_time,
                    csv_root=csv_root,
                    show_plot=False,
                    image_root=image_root,
                    tracking_root=tracking_root,
                )
                stats = _compute_trade_stats_from_df(algo_df)
                results.append({
                    "FileName": fname,
                    "Date": target_date,
                    "final_P/L": stats["final_pl"],
                    "winning_trades": stats["winning_trades"],
                    "losing_trades": stats["losing_trades"],
                    "win_pct": stats["win_pct"],
                    "Captured 100 points": stats.get("captured_100", "No"),
                    "p/l high": stats.get("pl_high", 0.0),
                    "p/l liquidation trade": stats.get("liquidation_trade_pl", None),
                })
            except Exception as exc:
                print(f"RunAllDays: error processing {fname}: {exc}")

        if results:
            # Write results CSV to the period folder
            df_res = pd.DataFrame(results)
            csv_simple = os.path.join(image_root, "RunAllDayscsv.csv")
            df_res.to_csv(csv_simple, index=False)
            print(f"Simple CSV summary written to: {csv_simple}")

    # ------------------------------------------------------------------ #
    # Combined 9:30-11:30 section                                          #
    # Runs three expanding windows over the merged data folder and writes  #
    # a RunAllDays_results.xlsx + per-day images into each sub-folder.    #
    # ------------------------------------------------------------------ #
    combined_csv_root = os.path.join(data_root, "CBOT_MINI_YM1_ByDate_930_1130")
    combined_parent = os.path.join(desktop_root, "TradingPics", "0930-1130")

    combined_windows = [
        ("09:30", "10:30", "0930-1030"),
        ("09:30", "11:00", "0930-1100"),
        ("09:30", "11:30", "0930-1130"),
    ]

    for start_time, end_time, sub_label in combined_windows:
        image_root = os.path.join(combined_parent, sub_label)
        tracking_root = os.path.join(desktop_root, "TradingTracking", "0930-1130", sub_label)
        os.makedirs(image_root, exist_ok=True)
        os.makedirs(tracking_root, exist_ok=True)

        pattern = os.path.join(combined_csv_root, "*.csv")
        files = sorted(glob.glob(pattern))
        if not files:
            print(f"RunAllDays: no CSV files found in {combined_csv_root}")
            continue

        results = []
        for fp in files:
            fname = os.path.basename(fp)
            target_date = _extract_date_from_filename(fname)
            if not target_date:
                print(f"RunAllDays: skipping {fname} (no parsable date)")
                continue

            print("\n" + "=" * 60)
            print(f"RunAllDays: processing {fname} (date={target_date}) [{start_time}-{end_time}]")
            print("=" * 60)

            try:
                algo_df = run_single_day(
                    target_date,
                    start_time,
                    end_time,
                    csv_root=combined_csv_root,
                    show_plot=False,
                    image_root=image_root,
                    tracking_root=tracking_root,
                )
                stats = _compute_trade_stats_from_df(algo_df)
                results.append({
                    "FileName": fname,
                    "Date": target_date,
                    "final_P/L": stats["final_pl"],
                    "winning_trades": stats["winning_trades"],
                    "losing_trades": stats["losing_trades"],
                    "win_pct": stats["win_pct"],
                    "Captured 100 points": stats.get("captured_100", "No"),
                    "p/l high": stats.get("pl_high", 0.0),
                    "p/l liquidation trade": stats.get("liquidation_trade_pl", None),
                })
            except Exception as exc:
                print(f"RunAllDays: error processing {fname}: {exc}")

        if results:
            _write_run_all_days_results(results, out_dir=image_root)
            print(f"Results for [{start_time}-{end_time}] written to: {image_root}")
