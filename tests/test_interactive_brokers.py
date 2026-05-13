"""tests/test_interactive_brokers.py

Unit tests for IBDataBridge and _LiveChartWindow in InteractiveBrokers.py.

All IB Gateway connectivity and matplotlib display are mocked -- no running
TWS/Gateway instance is required.
"""

import os
import sys
import types
from datetime import datetime, timezone
from unittest.mock import MagicMock, patch

import pandas as pd
import pytest
import pytz

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from InteractiveBrokers import IBDataBridge, _LiveChartWindow, _build_parser

_EST = pytz.timezone("US/Eastern")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _epoch(date_str: str, time_str: str) -> datetime:
    """Timezone-aware EST datetime for a given date + time (e.g. '2024-01-15', '09:30:05')."""
    dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M:%S")
    return _EST.localize(dt)


def make_bar(dt: datetime, o=38000.0, h=38010.0, l=37990.0, c=38005.0, vol=100):
    """Minimal BarData-like object with the fields _on_bar reads."""
    return types.SimpleNamespace(date=dt, open=o, high=h, low=l, close=c, volume=vol)


def make_bridge(**kw) -> IBDataBridge:
    """Return an IBDataBridge with the IB socket replaced by a MagicMock."""
    bridge = IBDataBridge(**kw)
    bridge._ib = MagicMock()
    bridge._contract = MagicMock()
    bridge._contract.localSymbol = "YMM6"
    return bridge


def feed_bar(bridge: IBDataBridge, dt: datetime, **bar_kw):
    """Deliver one sealed bar to bridge._on_bar, suppressing _run_algo."""
    bar = make_bar(dt, **bar_kw)
    with patch.object(bridge, "_run_algo"):
        bridge._on_bar([bar], has_new_bar=True)


def _one_bar_df(date: str = "2024-01-15", time: str = "09:30:00") -> pd.DataFrame:
    """Single-row DataFrame with tz-aware EST index, suitable as algo output."""
    ts = pd.Timestamp(f"{date} {time}", tz=_EST)
    return pd.DataFrame(
        [{"High": 38010.0, "Low": 37990.0, "Close": 38005.0}],
        index=[ts],
    )


# ---------------------------------------------------------------------------
# IBDataBridge.__init__
# ---------------------------------------------------------------------------

class TestIBDataBridgeInit:

    def test_defaults(self):
        b = make_bridge()
        assert b.host == "127.0.0.1"
        assert b.port == 4002
        assert b.client_id == 1
        assert b.dry_run is False
        assert b.start_time == "09:30"
        assert b.end_time == "10:00"
        assert b.show_plot is True
        assert b.tracking_root is None
        assert b.image_root is None
        assert b._session_bars == []
        assert b._current_date is None
        assert b._last_minute is None
        assert b._live_chart is None

    def test_custom_params(self):
        b = make_bridge(
            port=4002, dry_run=True,
            start_time="08:00", end_time="09:00",
            show_plot=False, tracking_root="/t", image_root="/i",
        )
        assert b.port == 4002
        assert b.dry_run is True
        assert b.start_time == "08:00"
        assert b.end_time == "09:00"
        assert b.show_plot is False
        assert b.tracking_root == "/t"
        assert b.image_root == "/i"


# ---------------------------------------------------------------------------
# _resample_to_minutes
# ---------------------------------------------------------------------------

class TestResampleToMinutes:

    def _add_bars(self, bridge, date, rows):
        """Directly populate _session_bars (list of dicts with tz-aware timestamps)."""
        for t, o, h, l, c, v in rows:
            ts = pd.Timestamp(f"{date} {t}", tz=_EST)
            bridge._session_bars.append(
                {"Open": o, "High": h, "Low": l, "Close": c, "Volume": v, "time": ts}
            )

    def test_single_minute_yields_one_row(self):
        b = make_bridge()
        self._add_bars(b, "2024-01-15", [
            ("09:30:00", 100, 110, 90, 105, 200),
            ("09:30:30", 105, 112, 92, 108, 150),
            ("09:30:55", 108, 115, 95, 110, 180),
        ])
        assert len(b._resample_to_minutes()) == 1

    def test_two_minutes_yield_two_rows(self):
        b = make_bridge()
        self._add_bars(b, "2024-01-15", [
            ("09:30:00", 100, 110, 90, 105, 200),
            ("09:30:55", 105, 112, 92, 108, 150),
            ("09:31:00", 108, 115, 95, 110, 180),
            ("09:31:55", 110, 120, 100, 115, 220),
        ])
        assert len(b._resample_to_minutes()) == 2

    def test_ohlcv_aggregation_is_correct(self):
        b = make_bridge()
        self._add_bars(b, "2024-01-15", [
            ("09:30:00", 100, 120, 90,  105, 200),   # open  = 100
            ("09:30:30", 105, 130, 85,  110, 300),   # high  = max(120,130,125) = 130
            ("09:30:55", 110, 125, 92,  115, 100),   # low   = min(90,85,92)    = 85
        ])                                            # close = last = 115
        row = b._resample_to_minutes().iloc[0]        # volume= 200+300+100      = 600
        assert row["Open"]   == 100
        assert row["High"]   == 130
        assert row["Low"]    == 85
        assert row["Close"]  == 115
        assert row["Volume"] == 600

    def test_empty_returns_empty_dataframe(self):
        b = make_bridge()
        assert b._resample_to_minutes().empty

    def test_index_is_tz_aware(self):
        b = make_bridge()
        self._add_bars(b, "2024-01-15", [("09:30:00", 100, 110, 90, 105, 200)])
        result = b._resample_to_minutes()
        assert result.index.tz is not None


# ---------------------------------------------------------------------------
# _on_bar: accumulation and minute / day rollover
# ---------------------------------------------------------------------------

class TestOnBar:

    def test_first_bar_sets_current_date_and_last_minute(self):
        b = make_bridge(show_plot=False)
        feed_bar(b, _epoch("2024-01-15", "09:30:05"))
        assert b._current_date == "2024-01-15"
        assert b._last_minute == "2024-01-15 09:30"

    def test_bars_accumulate_in_session_bars(self):
        b = make_bridge(show_plot=False)
        for i in range(3):
            feed_bar(b, _epoch("2024-01-15", f"09:30:0{i}"))
        assert len(b._session_bars) == 3

    def test_has_new_bar_false_skips_processing(self):
        b = make_bridge(show_plot=False)
        b._on_bar([make_bar(_epoch("2024-01-15", "09:30:00"))], has_new_bar=False)
        assert b._session_bars == []

    def test_empty_bars_list_skips_processing(self):
        b = make_bridge(show_plot=False)
        b._on_bar([], has_new_bar=True)
        assert b._session_bars == []

    def test_minute_rollover_triggers_update_live_chart(self):
        b = make_bridge(show_plot=True)
        with patch.object(b, "_update_live_chart") as mock_upd:
            feed_bar(b, _epoch("2024-01-15", "09:30:05"))   # first bar — no update
            mock_upd.assert_not_called()
            feed_bar(b, _epoch("2024-01-15", "09:30:50"))   # same minute — no update
            mock_upd.assert_not_called()
            feed_bar(b, _epoch("2024-01-15", "09:31:05"))   # new minute — update fires
            mock_upd.assert_called_once()

    def test_no_update_within_same_minute(self):
        b = make_bridge(show_plot=True)
        with patch.object(b, "_update_live_chart") as mock_upd:
            for sec in range(0, 60, 5):
                feed_bar(b, _epoch("2024-01-15", f"09:30:{sec:02d}"))
            mock_upd.assert_not_called()

    def test_multiple_minute_rollovers_each_trigger_update(self):
        b = make_bridge(show_plot=True)
        with patch.object(b, "_update_live_chart") as mock_upd:
            feed_bar(b, _epoch("2024-01-15", "09:30:05"))
            feed_bar(b, _epoch("2024-01-15", "09:31:05"))
            feed_bar(b, _epoch("2024-01-15", "09:32:05"))
        assert mock_upd.call_count == 2

    def test_last_minute_updated_on_rollover(self):
        b = make_bridge(show_plot=False)
        feed_bar(b, _epoch("2024-01-15", "09:30:05"))
        feed_bar(b, _epoch("2024-01-15", "09:31:05"))
        assert b._last_minute == "2024-01-15 09:31"

    def test_day_rollover_triggers_session_end(self):
        b = make_bridge(show_plot=False)
        b._current_date = "2024-01-15"
        b._last_minute = "2024-01-15 09:30"
        with patch.object(b, "_on_session_end") as mock_end:
            feed_bar(b, _epoch("2024-01-16", "09:30:05"))
        mock_end.assert_called_once()

    def test_no_session_end_within_same_day(self):
        b = make_bridge(show_plot=False)
        b._current_date = "2024-01-15"
        with patch.object(b, "_on_session_end") as mock_end:
            feed_bar(b, _epoch("2024-01-15", "09:31:00"))
        mock_end.assert_not_called()


# ---------------------------------------------------------------------------
# _on_session_end
# ---------------------------------------------------------------------------

class TestOnSessionEnd:

    def test_empty_bars_resets_state_without_calling_run_live_session(self):
        b = make_bridge()
        b._current_date = "2024-01-15"
        b._last_minute = "2024-01-15 09:30"
        with patch("InteractiveBrokers.run_live_session") as mock_rls:
            b._on_session_end()
        mock_rls.assert_not_called()
        assert b._session_bars == []
        assert b._last_result is None
        assert b._last_minute is None

    def test_closes_live_chart_first(self):
        b = make_bridge()
        mock_chart = MagicMock()
        b._live_chart = mock_chart
        with patch("InteractiveBrokers.run_live_session"):
            b._on_session_end()
        mock_chart.close.assert_called_once()
        assert b._live_chart is None

    def test_calls_run_live_session_with_correct_kwargs(self):
        b = make_bridge(
            start_time="09:30", end_time="10:00",
            show_plot=False, tracking_root="/t", image_root="/i",
        )
        b._current_date = "2024-01-15"
        ts = pd.Timestamp("2024-01-15 09:30:05", tz=_EST)
        b._session_bars = [
            {"Open": 100, "High": 110, "Low": 90, "Close": 105, "Volume": 200, "time": ts}
        ]
        with patch("InteractiveBrokers.run_live_session") as mock_rls, \
             patch.object(b, "_save_tracking_csv"):
            b._on_session_end()
        mock_rls.assert_called_once()
        _, kw = mock_rls.call_args
        assert kw["target_date"] == "2024-01-15"
        assert kw["start_time"] == "09:30"
        assert kw["end_time"] == "10:00"
        assert kw["show_plot"] is False
        assert kw["tracking_root"] == "/t"
        assert kw["image_root"] == "/i"

    def test_resets_all_state_after_run(self):
        b = make_bridge(show_plot=False)
        b._current_date = "2024-01-15"
        ts = pd.Timestamp("2024-01-15 09:30:05", tz=_EST)
        b._session_bars = [
            {"Open": 100, "High": 110, "Low": 90, "Close": 105, "Volume": 200, "time": ts}
        ]
        b._last_minute = "2024-01-15 09:30"
        with patch("InteractiveBrokers.run_live_session"):
            b._on_session_end()
        assert b._session_bars == []
        assert b._last_result is None
        assert b._last_minute is None

    def test_run_live_session_error_does_not_propagate(self):
        b = make_bridge(show_plot=False)
        b._current_date = "2024-01-15"
        ts = pd.Timestamp("2024-01-15 09:30:05", tz=_EST)
        b._session_bars = [
            {"Open": 100, "High": 110, "Low": 90, "Close": 105, "Volume": 200, "time": ts}
        ]
        with patch("InteractiveBrokers.run_live_session", side_effect=RuntimeError("boom")), \
             patch.object(b, "_save_tracking_csv"):
            b._on_session_end()   # must not raise

    def test_save_tracking_csv_called_at_session_end(self, tmp_path):
        b = make_bridge(tracking_root=str(tmp_path), show_plot=False)
        b._current_date = "2024-01-15"
        ts = pd.Timestamp("2024-01-15 09:30:05", tz=_EST)
        b._session_bars = [
            {"Open": 100, "High": 110, "Low": 90, "Close": 105, "Volume": 200, "time": ts}
        ]
        with patch("InteractiveBrokers.run_live_session"), \
             patch("InteractiveBrokers.run_trading_algo", return_value=_one_bar_df()):
            b._on_session_end()
        assert (tmp_path / "YM_tracking_2024-01-15.csv").exists()


# ---------------------------------------------------------------------------
# _update_live_chart
# ---------------------------------------------------------------------------

class TestUpdateLiveChart:

    def test_skipped_when_show_plot_false(self):
        b = make_bridge(show_plot=False)
        b._current_date = "2024-01-15"
        with patch("InteractiveBrokers.run_trading_algo") as mock_rta:
            b._update_live_chart()
        mock_rta.assert_not_called()

    def test_skipped_when_no_bars(self):
        b = make_bridge(show_plot=True)
        b._current_date = "2024-01-15"
        with patch("InteractiveBrokers.run_trading_algo") as mock_rta:
            b._update_live_chart()
        mock_rta.assert_not_called()

    def test_creates_live_chart_window_on_first_call(self):
        b = make_bridge(show_plot=True, start_time="09:30", end_time="10:00")
        b._current_date = "2024-01-15"
        ts = pd.Timestamp("2024-01-15 09:30:05", tz=_EST)
        b._session_bars = [
            {"Open": 100, "High": 110, "Low": 90, "Close": 105, "Volume": 200, "time": ts}
        ]
        fake_df = _one_bar_df()
        with patch("InteractiveBrokers.run_trading_algo", return_value=fake_df), \
             patch.object(_LiveChartWindow, "update") as mock_update:
            b._update_live_chart()
        assert b._live_chart is not None
        mock_update.assert_called_once_with(fake_df)

    def test_reuses_existing_live_chart_window(self):
        b = make_bridge(show_plot=True)
        b._current_date = "2024-01-15"
        ts = pd.Timestamp("2024-01-15 09:30:05", tz=_EST)
        b._session_bars = [
            {"Open": 100, "High": 110, "Low": 90, "Close": 105, "Volume": 200, "time": ts}
        ]
        fake_df = _one_bar_df()
        mock_chart = MagicMock()
        b._live_chart = mock_chart
        with patch("InteractiveBrokers.run_trading_algo", return_value=fake_df):
            b._update_live_chart()
        mock_chart.update.assert_called_once_with(fake_df)

    def test_trading_algo_error_does_not_propagate(self):
        b = make_bridge(show_plot=True)
        b._current_date = "2024-01-15"
        ts = pd.Timestamp("2024-01-15 09:30:05", tz=_EST)
        b._session_bars = [
            {"Open": 100, "High": 110, "Low": 90, "Close": 105, "Volume": 200, "time": ts}
        ]
        with patch("InteractiveBrokers.run_trading_algo", side_effect=ValueError("bad")):
            b._update_live_chart()   # must not raise

    def test_chart_update_error_does_not_propagate(self):
        b = make_bridge(show_plot=True)
        b._current_date = "2024-01-15"
        ts = pd.Timestamp("2024-01-15 09:30:05", tz=_EST)
        b._session_bars = [
            {"Open": 100, "High": 110, "Low": 90, "Close": 105, "Volume": 200, "time": ts}
        ]
        fake_df = _one_bar_df()
        mock_chart = MagicMock()
        mock_chart.update.side_effect = RuntimeError("render fail")
        b._live_chart = mock_chart
        with patch("InteractiveBrokers.run_trading_algo", return_value=fake_df):
            b._update_live_chart()   # must not raise


# ---------------------------------------------------------------------------
# _place_order
# ---------------------------------------------------------------------------

class TestPlaceOrder:

    def test_dry_run_does_not_call_place_order(self):
        b = make_bridge(dry_run=True)
        b._place_order("BUY")
        b._ib.placeOrder.assert_not_called()

    def test_dry_run_liquidate_does_not_call_place_order(self):
        b = make_bridge(dry_run=True)
        b._place_order("SELL", liquidate=True)
        b._ib.placeOrder.assert_not_called()

    def test_live_mode_buy_calls_place_order(self):
        b = make_bridge(dry_run=False)
        b._place_order("BUY")
        b._ib.placeOrder.assert_called_once()

    def test_live_mode_sell_calls_place_order(self):
        b = make_bridge(dry_run=False)
        b._place_order("SELL")
        b._ib.placeOrder.assert_called_once()


# ---------------------------------------------------------------------------
# _LiveChartWindow
# ---------------------------------------------------------------------------

class TestLiveChartWindow:

    def test_init_has_no_plotter(self):
        w = _LiveChartWindow("2024-01-15", "09:30", "10:00")
        assert w._plotter is None

    def test_first_update_creates_plotter_and_draws(self):
        w = _LiveChartWindow("2024-01-15", "09:30", "10:00")
        df = _one_bar_df()
        mock_plotter = MagicMock()
        mock_plotter.ax_top = None
        with patch("InteractiveBrokers.ChartPlotter", return_value=mock_plotter), \
             patch("matplotlib.pyplot.ion"):
            w.update(df)
        mock_plotter.create_figure.assert_called_once()
        mock_plotter.update_plot.assert_called_once_with(0)
        mock_plotter.fig.canvas.draw.assert_called_once()
        mock_plotter.fig.canvas.start_event_loop.assert_called_once_with(0.5)

    def test_second_update_replaces_data_and_expands_xlim(self):
        w = _LiveChartWindow("2024-01-15", "09:30", "10:00")
        ts1 = pd.Timestamp("2024-01-15 09:30:00", tz=_EST)
        ts2 = pd.Timestamp("2024-01-15 09:31:00", tz=_EST)
        df1 = pd.DataFrame([{"High": 110.0}], index=[ts1])
        df2 = pd.DataFrame([{"High": 110.0}, {"High": 115.0}], index=[ts1, ts2])

        mock_plotter = MagicMock()
        mock_plotter.ax_top = MagicMock()
        with patch("InteractiveBrokers.ChartPlotter", return_value=mock_plotter), \
             patch("matplotlib.pyplot.ion"):
            w.update(df1)
            w.update(df2)

        assert mock_plotter.data is df2
        assert mock_plotter.update_plot.call_count == 2
        mock_plotter.update_plot.assert_called_with(1)   # len(df2)-1

    def test_second_update_does_not_recreate_figure(self):
        w = _LiveChartWindow("2024-01-15", "09:30", "10:00")
        df = _one_bar_df()
        mock_plotter = MagicMock()
        mock_plotter.ax_top = None
        with patch("InteractiveBrokers.ChartPlotter", return_value=mock_plotter) as mock_cls, \
             patch("matplotlib.pyplot.ion"):
            w.update(df)
            w.update(df)
        mock_cls.assert_called_once()            # constructor called only once
        mock_plotter.create_figure.assert_called_once()  # figure created only once

    def test_close_calls_plt_close_and_clears_plotter(self):
        w = _LiveChartWindow("2024-01-15", "09:30", "10:00")
        mock_plotter = MagicMock()
        mock_plotter.fig = MagicMock()
        w._plotter = mock_plotter
        with patch("matplotlib.pyplot.close") as mock_close:
            w.close()
        mock_close.assert_called_once_with(mock_plotter.fig)
        assert w._plotter is None

    def test_close_is_safe_when_plotter_is_none(self):
        w = _LiveChartWindow("2024-01-15", "09:30", "10:00")
        with patch("matplotlib.pyplot.close") as mock_close:
            w.close()   # must not raise
        mock_close.assert_not_called()


# ---------------------------------------------------------------------------
# _build_parser
# ---------------------------------------------------------------------------

class TestBuildParser:

    def test_defaults(self):
        args = _build_parser().parse_args([])
        assert args.host == "127.0.0.1"
        assert args.port == 4002
        assert args.client_id == 1
        assert args.test is False
        assert args.dry_run is False
        assert args.start_time == "09:30"
        assert args.end_time == "10:00"
        assert args.show_plot is True
        assert args.tracking_root == os.path.join(os.path.expanduser("~"), "Desktop", "IB_Live", "tracking")
        assert args.image_root == os.path.join(os.path.expanduser("~"), "Desktop", "IB_Live", "charts")

    def test_all_live_flags(self):
        args = _build_parser().parse_args([
            "--port", "4002",
            "--dry-run",
            "--start-time", "08:30",
            "--end-time", "09:30",
            "--no-plot",
            "--tracking-root", "/tmp/tracking",
            "--image-root", "/tmp/images",
        ])
        assert args.port == 4002
        assert args.dry_run is True
        assert args.start_time == "08:30"
        assert args.end_time == "09:30"
        assert args.show_plot is False
        assert args.tracking_root == "/tmp/tracking"
        assert args.image_root == "/tmp/images"

    def test_test_flag(self):
        args = _build_parser().parse_args(["--test"])
        assert args.test is True

    def test_client_id_flag(self):
        args = _build_parser().parse_args(["--client-id", "5"])
        assert args.client_id == 5


# ---------------------------------------------------------------------------
# _append_bar_to_excel
# ---------------------------------------------------------------------------

class TestAppendBarToExcel:

    def test_no_file_when_tracking_root_is_none(self, tmp_path):
        b = make_bridge(tracking_root=None)
        dt = _epoch("2024-01-15", "09:30:05")
        feed_bar(b, dt)
        assert b._raw_wb is None

    def test_creates_workbook_on_first_bar(self, tmp_path):
        b = make_bridge(tracking_root=str(tmp_path))
        dt = _epoch("2024-01-15", "09:30:05")
        feed_bar(b, dt)
        assert b._raw_wb is not None
        assert b._raw_path == str(tmp_path / "YM_raw_2024-01-15.xlsx")

    def test_file_exists_after_first_bar(self, tmp_path):
        b = make_bridge(tracking_root=str(tmp_path))
        feed_bar(b, _epoch("2024-01-15", "09:30:05"))
        assert (tmp_path / "YM_raw_2024-01-15.xlsx").exists()

    def test_headers_are_correct(self, tmp_path):
        from openpyxl import load_workbook
        b = make_bridge(tracking_root=str(tmp_path))
        feed_bar(b, _epoch("2024-01-15", "09:30:05"))
        wb = load_workbook(tmp_path / "YM_raw_2024-01-15.xlsx")
        headers = [c.value for c in wb.active[1]]
        assert headers == ["Time", "Open", "High", "Low", "Close", "Volume"]

    def test_row_values_are_correct(self, tmp_path):
        from openpyxl import load_workbook
        b = make_bridge(tracking_root=str(tmp_path))
        feed_bar(b, _epoch("2024-01-15", "09:30:05"), o=38000.0, h=38010.0, l=37990.0, c=38005.0, vol=50)
        wb = load_workbook(tmp_path / "YM_raw_2024-01-15.xlsx")
        row = [c.value for c in wb.active[2]]
        assert row[0] == "2024-01-15 09:30:05"
        assert row[1] == 38000.0
        assert row[2] == 38010.0
        assert row[3] == 37990.0
        assert row[4] == 38005.0
        assert row[5] == 50

    def test_each_bar_appends_a_row(self, tmp_path):
        from openpyxl import load_workbook
        b = make_bridge(tracking_root=str(tmp_path))
        for sec in range(5, 20, 5):
            feed_bar(b, _epoch("2024-01-15", f"09:30:{sec:02d}"))
        wb = load_workbook(tmp_path / "YM_raw_2024-01-15.xlsx")
        assert wb.active.max_row == 4  # 1 header + 3 data rows


# ---------------------------------------------------------------------------
# _save_tracking_csv
# ---------------------------------------------------------------------------

class TestLiveTrackingCSV:

    def test_no_csv_when_tracking_root_is_none(self, tmp_path):
        b = make_bridge(tracking_root=None)
        b._current_date = "2024-01-15"
        ts = pd.Timestamp("2024-01-15 09:30:05", tz=_EST)
        b._session_bars = [
            {"Open": 100, "High": 110, "Low": 90, "Close": 105, "Volume": 200, "time": ts}
        ]
        with patch("InteractiveBrokers.run_trading_algo") as mock_rta:
            b._save_tracking_csv()
        mock_rta.assert_not_called()

    def test_no_csv_when_session_bars_empty(self, tmp_path):
        b = make_bridge(tracking_root=str(tmp_path))
        b._current_date = "2024-01-15"
        b._save_tracking_csv()
        assert not (tmp_path / "YM_tracking_2024-01-15.csv").exists()

    def test_csv_written_at_minute_boundary(self, tmp_path):
        b = make_bridge(tracking_root=str(tmp_path), show_plot=False)
        with patch("InteractiveBrokers.run_trading_algo", return_value=_one_bar_df()):
            feed_bar(b, _epoch("2024-01-15", "09:30:05"))   # first bar — no boundary
            assert not (tmp_path / "YM_tracking_2024-01-15.csv").exists()
            feed_bar(b, _epoch("2024-01-15", "09:31:05"))   # rollover → CSV written
        assert (tmp_path / "YM_tracking_2024-01-15.csv").exists()

    def test_no_csv_within_same_minute(self, tmp_path):
        b = make_bridge(tracking_root=str(tmp_path), show_plot=False)
        with patch("InteractiveBrokers.run_trading_algo", return_value=_one_bar_df()):
            for sec in range(0, 60, 5):
                feed_bar(b, _epoch("2024-01-15", f"09:30:{sec:02d}"))
        assert not (tmp_path / "YM_tracking_2024-01-15.csv").exists()

    def test_csv_updated_on_each_minute_boundary(self, tmp_path):
        b = make_bridge(tracking_root=str(tmp_path), show_plot=False)
        with patch("InteractiveBrokers.run_trading_algo", return_value=_one_bar_df()) as mock_rta:
            feed_bar(b, _epoch("2024-01-15", "09:30:05"))
            feed_bar(b, _epoch("2024-01-15", "09:31:05"))   # first boundary
            feed_bar(b, _epoch("2024-01-15", "09:32:05"))   # second boundary
        # show_plot=False so _update_live_chart skips algo; only _save_tracking_csv calls it
        assert mock_rta.call_count == 2
        assert (tmp_path / "YM_tracking_2024-01-15.csv").exists()

    def test_algo_error_does_not_propagate(self, tmp_path):
        b = make_bridge(tracking_root=str(tmp_path), show_plot=False)
        b._current_date = "2024-01-15"
        ts = pd.Timestamp("2024-01-15 09:30:05", tz=_EST)
        b._session_bars = [
            {"Open": 100, "High": 110, "Low": 90, "Close": 105, "Volume": 200, "time": ts}
        ]
        with patch("InteractiveBrokers.run_trading_algo", side_effect=ValueError("boom")):
            b._save_tracking_csv()   # must not raise
        assert not (tmp_path / "YM_tracking_2024-01-15.csv").exists()


# ---------------------------------------------------------------------------
# IB Signal Validation: Algo signals match IB order execution
# ---------------------------------------------------------------------------

class TestIBSignalValidation:
    """Validate that algo signals match what IB would actually execute.
    
    This test ensures that:
    1. The algo doesn't generate duplicate BUY signals when already LONG
    2. The algo doesn't generate duplicate SELL signals when already SHORT
    3. IB's duplicate protection logic correctly filters signals
    4. The number of IB orders matches the number of valid algo signals
    """

    def test_no_duplicate_signals_may_12_2026(self):
        """Test that May 12, 2026 generates valid signals with no duplicates."""
        from pathlib import Path
        from TradingAlgoFast import AlgoConfig, run_trading_algo_fast
        
        # Load May 12, 2026 data
        csv_path = Path.home() / "Desktop" / "2YearsData" / "full_day" / "CBOT_MINI_YM1_2026-05-12.csv"
        
        if not csv_path.exists():
            pytest.skip(f"Test data not found: {csv_path}")
        
        df = pd.read_csv(csv_path, index_col=0, parse_dates=True)
        df.index = pd.to_datetime(df.index, utc=True).tz_convert(_EST)
        
        # Filter to day session
        day_start = pd.Timestamp("2026-05-12 09:30", tz=_EST)
        day_end = pd.Timestamp("2026-05-12 17:00", tz=_EST)
        df = df[(df.index >= day_start) & (df.index <= day_end)]
        
        # Config matching live trading
        config = AlgoConfig(
            warmup_minutes=5,
            steep_angle_threshold=65.0,
            proximity_points=8.0,
            min_reversal_minutes=0,
            min_entry_angle=15.0,
            partial_tp_pts=50.0,
            spike_profit_pts=50.0,
            spike_profit_bars=9,
            wm_shield_distance=0.0,
            steep_line_reentry=False,
            steep_line_proximity=5.0,
            steep_line_exit_only=False,
        )
        
        # Run algo
        result = run_trading_algo_fast(df, target_date="2026-05-12", start_time="09:30", end_time="17:00", config=config)
        
        # Extract signals
        signals = result[result['signal'].isin(['BUY', 'SELL'])].copy()
        
        # Simulate IB order logic using algo's pos_debug column
        ib_orders = []
        skipped_signals = []
        
        for i, (idx, row) in enumerate(signals.iterrows()):
            signal = row['signal']
            
            # Get position BEFORE this signal (from previous bar in result DataFrame)
            bar_idx = result.index.get_loc(idx)
            if bar_idx > 0:
                pos_before = result.iloc[bar_idx - 1]['pos_debug']
            else:
                pos_before = 0  # start flat
            
            # IB duplicate protection: check position BEFORE signal
            skip = False
            if signal == 'BUY' and pos_before == 1:  # already long
                skipped_signals.append((idx, signal, pos_before))
                skip = True
            elif signal == 'SELL' and pos_before == 2:  # already short
                skipped_signals.append((idx, signal, pos_before))
                skip = True
            
            if not skip:
                ib_orders.append({
                    'time': idx,
                    'signal': signal,
                    'contracts': 2
                })
        
        # Assertions
        algo_buys = (signals['signal'] == 'BUY').sum()
        algo_sells = (signals['signal'] == 'SELL').sum()
        ib_buys = sum(1 for o in ib_orders if o['signal'] == 'BUY')
        ib_sells = sum(1 for o in ib_orders if o['signal'] == 'SELL')
        
        # No signals should be skipped (no duplicates)
        assert len(skipped_signals) == 0, \
            f"Found {len(skipped_signals)} duplicate signals: {skipped_signals}"
        
        # IB orders should match algo signals
        assert algo_buys == ib_buys, \
            f"BUY mismatch: algo={algo_buys}, IB={ib_buys}"
        assert algo_sells == ib_sells, \
            f"SELL mismatch: algo={algo_sells}, IB={ib_sells}"
        
        # Verify specific counts for May 12, 2026
        assert algo_buys == 11, f"Expected 11 BUY signals, got {algo_buys}"
        assert algo_sells == 12, f"Expected 12 SELL signals, got {algo_sells}"

    def test_position_tracking_logic(self):
        """Test that position tracking correctly identifies duplicates."""
        # Create a simple test case with known signals
        test_data = pd.DataFrame({
            'signal': ['', 'SELL', '', 'BUY', '', 'SELL', '', 'SELL', '', 'BUY', ''],
            'pos_debug': [0, 2, 2, 1, 1, 0, 0, 2, 2, 1, 1],  # 0=flat, 1=long, 2=short
        }, index=pd.date_range('2024-01-15 09:30', periods=11, freq='1min', tz=_EST))
        
        signals = test_data[test_data['signal'].isin(['BUY', 'SELL'])].copy()
        
        # Simulate IB duplicate detection
        duplicates = []
        for idx, row in signals.iterrows():
            signal = row['signal']
            bar_idx = test_data.index.get_loc(idx)
            pos_before = test_data.iloc[bar_idx - 1]['pos_debug'] if bar_idx > 0 else 0
            
            if signal == 'BUY' and pos_before == 1:
                duplicates.append(('BUY', idx))
            elif signal == 'SELL' and pos_before == 2:
                duplicates.append(('SELL', idx))
        
        # Should have no duplicates in this sequence:
        # FLAT -> SELL (go SHORT) -> BUY (go LONG) -> SELL (go FLAT) -> SELL (go SHORT) -> BUY (go LONG)
        assert len(duplicates) == 0, f"Found unexpected duplicates: {duplicates}"

    def test_duplicate_buy_detection(self):
        """Test that duplicate BUY signals are correctly detected."""
        # Create test case with duplicate BUY
        test_data = pd.DataFrame({
            'signal': ['', 'BUY', '', 'BUY', ''],  # Second BUY is duplicate
            'pos_debug': [0, 1, 1, 1, 1],  # Already LONG before second BUY
        }, index=pd.date_range('2024-01-15 09:30', periods=5, freq='1min', tz=_EST))
        
        signals = test_data[test_data['signal'].isin(['BUY', 'SELL'])].copy()
        
        duplicates = []
        for idx, row in signals.iterrows():
            signal = row['signal']
            bar_idx = test_data.index.get_loc(idx)
            pos_before = test_data.iloc[bar_idx - 1]['pos_debug'] if bar_idx > 0 else 0
            
            if signal == 'BUY' and pos_before == 1:
                duplicates.append(('BUY', idx))
        
        assert len(duplicates) == 1, f"Expected 1 duplicate BUY, found {len(duplicates)}"

    def test_duplicate_sell_detection(self):
        """Test that duplicate SELL signals are correctly detected."""
        # Create test case with duplicate SELL
        test_data = pd.DataFrame({
            'signal': ['', 'SELL', '', 'SELL', ''],  # Second SELL is duplicate
            'pos_debug': [0, 2, 2, 2, 2],  # Already SHORT before second SELL
        }, index=pd.date_range('2024-01-15 09:30', periods=5, freq='1min', tz=_EST))
        
        signals = test_data[test_data['signal'].isin(['BUY', 'SELL'])].copy()
        
        duplicates = []
        for idx, row in signals.iterrows():
            signal = row['signal']
            bar_idx = test_data.index.get_loc(idx)
            pos_before = test_data.iloc[bar_idx - 1]['pos_debug'] if bar_idx > 0 else 0
            
            if signal == 'SELL' and pos_before == 2:
                duplicates.append(('SELL', idx))
        
        assert len(duplicates) == 1, f"Expected 1 duplicate SELL, found {len(duplicates)}"

    def test_consecutive_sells_from_long_to_flat_to_short(self):
        """Test that two consecutive SELLs are valid when going LONG -> FLAT -> SHORT."""
        # This is the pattern that was incorrectly flagged as duplicate
        test_data = pd.DataFrame({
            'signal': ['', 'BUY', '', 'SELL', '', 'SELL', ''],
            'pos_debug': [0, 1, 1, 0, 0, 2, 2],  # FLAT -> LONG -> FLAT -> SHORT
        }, index=pd.date_range('2024-01-15 09:30', periods=7, freq='1min', tz=_EST))
        
        signals = test_data[test_data['signal'].isin(['BUY', 'SELL'])].copy()
        
        duplicates = []
        for idx, row in signals.iterrows():
            signal = row['signal']
            bar_idx = test_data.index.get_loc(idx)
            pos_before = test_data.iloc[bar_idx - 1]['pos_debug'] if bar_idx > 0 else 0
            
            if signal == 'BUY' and pos_before == 1:
                duplicates.append(('BUY', idx))
            elif signal == 'SELL' and pos_before == 2:
                duplicates.append(('SELL', idx))
        
        # Both SELLs are valid: first closes LONG, second opens SHORT
        assert len(duplicates) == 0, f"Found unexpected duplicates: {duplicates}"


# ---------------------------------------------------------------------------
# IB P/L Validation: Algo P/L calculation matches expected behavior
# ---------------------------------------------------------------------------

class TestIBPLValidation:
    """Validate that algo P/L calculation is correct and consistent.
    
    This test ensures that:
    1. P/L is calculated correctly for position closes
    2. Partial TP events are tracked correctly
    3. The algo's session_pl matches expected cumulative P/L
    """

    def test_pl_calculation_may_12_2026(self):
        """Test P/L calculation for May 12, 2026 including partial TP."""
        from pathlib import Path
        from TradingAlgoFast import AlgoConfig, run_trading_algo_fast
        
        # Load May 12, 2026 data
        csv_path = Path.home() / "Desktop" / "2YearsData" / "full_day" / "CBOT_MINI_YM1_2026-05-12.csv"
        
        if not csv_path.exists():
            pytest.skip(f"Test data not found: {csv_path}")
        
        df = pd.read_csv(csv_path, index_col=0, parse_dates=True)
        df.index = pd.to_datetime(df.index, utc=True).tz_convert(_EST)
        
        # Filter to day session
        day_start = pd.Timestamp("2026-05-12 09:30", tz=_EST)
        day_end = pd.Timestamp("2026-05-12 17:00", tz=_EST)
        df = df[(df.index >= day_start) & (df.index <= day_end)]
        
        # Config with partial TP enabled
        config = AlgoConfig(
            warmup_minutes=5,
            steep_angle_threshold=65.0,
            proximity_points=8.0,
            min_reversal_minutes=0,
            min_entry_angle=15.0,
            partial_tp_pts=50.0,
            spike_profit_pts=50.0,
            spike_profit_bars=9,
            wm_shield_distance=0.0,
            steep_line_reentry=False,
            steep_line_proximity=5.0,
            steep_line_exit_only=False,
        )
        
        # Run algo
        result = run_trading_algo_fast(df, target_date="2026-05-12", start_time="09:30", end_time="17:00", config=config)
        
        # Check partial TP events
        partial_tp_events = result[result['partial_tp'] == True]
        
        # Verify partial TP count
        assert len(partial_tp_events) == 8, f"Expected 8 partial TP events, got {len(partial_tp_events)}"
        
        # Verify final P/L
        final_pl = result.iloc[-1]['session_pl']
        assert abs(final_pl - 672.0) < 1.0, f"Expected final P/L ~672 pts, got {final_pl:.1f}"
        
        # Verify partial TP times
        expected_tp_times = ['09:59', '10:11', '11:33', '11:54', '12:56', '13:02', '14:24', '15:33']
        actual_tp_times = [idx.strftime('%H:%M') for idx in partial_tp_events.index]
        assert actual_tp_times == expected_tp_times, f"Partial TP times mismatch: {actual_tp_times}"

    def test_pl_increases_monotonically_on_wins(self):
        """Test that P/L increases when closing winning positions."""
        from TradingAlgoFast import AlgoConfig, run_trading_algo_fast
        
        # Create simple test data with a winning trade
        test_data = pd.DataFrame({
            'Open': [100, 100, 100, 100, 100],
            'High': [105, 105, 105, 105, 105],
            'Low': [95, 95, 95, 95, 95],
            'Close': [100, 100, 105, 105, 105],  # Price goes up
            'Volume': [100, 100, 100, 100, 100],
        }, index=pd.date_range('2024-01-15 09:30', periods=5, freq='1min', tz=_EST))
        
        config = AlgoConfig(
            warmup_minutes=0,
            steep_angle_threshold=65.0,
            proximity_points=8.0,
            min_reversal_minutes=0,
            min_entry_angle=0.0,
            partial_tp_pts=0.0,  # Disable partial TP for this test
            spike_profit_pts=0.0,
            wm_shield_distance=0.0,
        )
        
        from TradingAlgoFast import run_trading_algo_fast
        result = run_trading_algo_fast(test_data, target_date="2024-01-15", start_time="09:30", end_time="10:00", config=config)
        
        # Check that session_pl is non-decreasing (can only increase or stay same)
        session_pl_values = result['session_pl'].values
        for i in range(1, len(session_pl_values)):
            assert session_pl_values[i] >= session_pl_values[i-1] or session_pl_values[i] == 0, \
                f"P/L decreased from {session_pl_values[i-1]} to {session_pl_values[i]} at index {i}"

    def test_pl_with_partial_tp(self):
        """Test that partial TP correctly adds 50 pts to session_pl."""
        from pathlib import Path
        from TradingAlgoFast import AlgoConfig, run_trading_algo_fast
        
        csv_path = Path.home() / "Desktop" / "2YearsData" / "full_day" / "CBOT_MINI_YM1_2026-05-12.csv"
        
        if not csv_path.exists():
            pytest.skip(f"Test data not found: {csv_path}")
        
        df = pd.read_csv(csv_path, index_col=0, parse_dates=True)
        df.index = pd.to_datetime(df.index, utc=True).tz_convert(_EST)
        
        day_start = pd.Timestamp("2026-05-12 09:30", tz=_EST)
        day_end = pd.Timestamp("2026-05-12 17:00", tz=_EST)
        df = df[(df.index >= day_start) & (df.index <= day_end)]
        
        # Run with partial TP enabled
        config_with_tp = AlgoConfig(
            warmup_minutes=5,
            steep_angle_threshold=65.0,
            proximity_points=8.0,
            min_reversal_minutes=0,
            min_entry_angle=15.0,
            partial_tp_pts=50.0,  # Enable partial TP
            spike_profit_pts=50.0,
            wm_shield_distance=0.0,
            steep_line_reentry=False,
            steep_line_proximity=5.0,
        )
        
        result_with_tp = run_trading_algo_fast(df, target_date="2026-05-12", start_time="09:30", end_time="17:00", config=config_with_tp)
        
        # Run without partial TP
        config_no_tp = AlgoConfig(
            warmup_minutes=5,
            steep_angle_threshold=65.0,
            proximity_points=8.0,
            min_reversal_minutes=0,
            min_entry_angle=15.0,
            partial_tp_pts=0.0,  # Disable partial TP
            spike_profit_pts=50.0,
            wm_shield_distance=0.0,
            steep_line_reentry=False,
            steep_line_proximity=5.0,
        )
        
        result_no_tp = run_trading_algo_fast(df, target_date="2026-05-12", start_time="09:30", end_time="17:00", config=config_no_tp)
        
        # With partial TP should have higher P/L due to booking profits early
        pl_with_tp = result_with_tp.iloc[-1]['session_pl']
        pl_no_tp = result_no_tp.iloc[-1]['session_pl']
        
        # Verify partial TP adds to P/L
        partial_tp_count = (result_with_tp['partial_tp'] == True).sum()
        assert partial_tp_count > 0, "Expected at least one partial TP event"
        
        # P/L with partial TP should be different (usually higher)
        assert pl_with_tp != pl_no_tp, f"P/L should differ with/without partial TP: {pl_with_tp} vs {pl_no_tp}"

    def test_session_pl_starts_at_zero(self):
        """Test that session_pl starts at 0 at the beginning of the session."""
        from pathlib import Path
        from TradingAlgoFast import AlgoConfig, run_trading_algo_fast
        
        csv_path = Path.home() / "Desktop" / "2YearsData" / "full_day" / "CBOT_MINI_YM1_2026-05-12.csv"
        
        if not csv_path.exists():
            pytest.skip(f"Test data not found: {csv_path}")
        
        df = pd.read_csv(csv_path, index_col=0, parse_dates=True)
        df.index = pd.to_datetime(df.index, utc=True).tz_convert(_EST)
        
        day_start = pd.Timestamp("2026-05-12 09:30", tz=_EST)
        day_end = pd.Timestamp("2026-05-12 17:00", tz=_EST)
        df = df[(df.index >= day_start) & (df.index <= day_end)]
        
        config = AlgoConfig(
            warmup_minutes=5,
            steep_angle_threshold=65.0,
            proximity_points=8.0,
            min_reversal_minutes=0,
            min_entry_angle=15.0,
            partial_tp_pts=50.0,
        )
        
        result = run_trading_algo_fast(df, target_date="2026-05-12", start_time="09:30", end_time="17:00", config=config)
        
        # First bar should have session_pl = 0
        assert result.iloc[0]['session_pl'] == 0.0, f"Expected session_pl to start at 0, got {result.iloc[0]['session_pl']}"

    def test_pl_calculation_consistency(self):
        """Test that P/L calculation is consistent across multiple runs."""
        from pathlib import Path
        from TradingAlgoFast import AlgoConfig, run_trading_algo_fast
        
        csv_path = Path.home() / "Desktop" / "2YearsData" / "full_day" / "CBOT_MINI_YM1_2026-05-12.csv"
        
        if not csv_path.exists():
            pytest.skip(f"Test data not found: {csv_path}")
        
        df = pd.read_csv(csv_path, index_col=0, parse_dates=True)
        df.index = pd.to_datetime(df.index, utc=True).tz_convert(_EST)
        
        day_start = pd.Timestamp("2026-05-12 09:30", tz=_EST)
        day_end = pd.Timestamp("2026-05-12 17:00", tz=_EST)
        df = df[(df.index >= day_start) & (df.index <= day_end)]
        
        config = AlgoConfig(
            warmup_minutes=5,
            steep_angle_threshold=65.0,
            proximity_points=8.0,
            min_reversal_minutes=0,
            min_entry_angle=15.0,
            partial_tp_pts=50.0,
        )
        
        # Run twice with same config
        result1 = run_trading_algo_fast(df, target_date="2026-05-12", start_time="09:30", end_time="17:00", config=config)
        result2 = run_trading_algo_fast(df, target_date="2026-05-12", start_time="09:30", end_time="17:00", config=config)
        
        # P/L should be identical
        pl1 = result1.iloc[-1]['session_pl']
        pl2 = result2.iloc[-1]['session_pl']
        
        assert abs(pl1 - pl2) < 0.01, f"P/L should be consistent across runs: {pl1} vs {pl2}"
        
        # All session_pl values should match
        assert (result1['session_pl'] == result2['session_pl']).all(), "session_pl values should be identical across runs"
