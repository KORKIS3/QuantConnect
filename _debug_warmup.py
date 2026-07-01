"""Debug: Read Fred's actual raw bar data for 06/23 and check what the algo sees."""
import pandas as pd, pytz, os
from TradingAlgoFast import AlgoConfig, run_trading_algo_fast
from openpyxl import load_workbook

EST = pytz.timezone("US/Eastern")

# Read the raw Excel file Fred saved
xlsx_path = os.path.expanduser("~/Desktop/IB_Live/tracking/YM_raw_DUO158495_2026-06-23.xlsx")
wb = load_workbook(xlsx_path, read_only=True)
ws = wb.active

rows = list(ws.iter_rows(min_row=2, values_only=True))
wb.close()

# Build DataFrame from raw 5-sec bars
df = pd.DataFrame(rows, columns=["time", "Open", "High", "Low", "Close", "Volume"])
df["time"] = pd.to_datetime(df["time"])
df = df.set_index("time")
df.index = df.index.tz_localize(EST)

print(f"Raw bars: {len(df)}, from {df.index[0]} to {df.index[-1]}")

# Filter to session start (09:30) like _resample_to_minutes does
session_start = pd.Timestamp("2026-06-23 09:30", tz=EST)
df_session = df[df.index >= session_start]
print(f"After session filter (>= 09:30): {len(df_session)} bars, first={df_session.index[0]}")

# Resample to 1-min bars (same as Fred's _resample_to_minutes)
minute_df = df_session.resample("1min").agg(
    Open=("Open", "first"),
    High=("High", "max"),
    Low=("Low", "min"),
    Close=("Close", "last"),
    Volume=("Volume", "sum"),
).dropna(subset=["Open"])

print(f"Minute bars: {len(minute_df)}, first={minute_df.index[0]}")
print(f"\nFirst 5 minute bars (Fred's live data):")
for idx, row in minute_df.head(5).iterrows():
    print(f"  {idx.strftime('%H:%M')}  O={row['Open']:.0f} H={row['High']:.0f} L={row['Low']:.0f} C={row['Close']:.0f}")

# Now run algo on Fred's actual live data
config = AlgoConfig(
    warmup_minutes=7, steep_angle_threshold=90.0, proximity_points=4.0,
    min_reversal_minutes=0, min_entry_angle=0.0, partial_tp_pts=50.0,
    spike_profit_pts=100.0, spike_profit_bars=9, wm_shield_distance=0.0,
    swing_anchor_threshold=10.0, num_contracts=2, cushion_points=0.0, limit_expiry_bars=5,
)

# Simulate what Fred sees at each minute boundary (growing window)
# At 09:33, Fred has bars 09:30-09:33. Run algo on that.
print("\n=== SIMULATING LIVE ALGO BAR-BY-BAR ===")
for end_min in range(4, 12):  # 09:34 through 09:41
    end_time = pd.Timestamp(f"2026-06-23 09:{30+end_min}:00", tz=EST)
    growing_df = minute_df[minute_df.index <= end_time]
    if len(growing_df) < 2:
        continue
    try:
        result = run_trading_algo_fast(growing_df, "2026-06-23", "09:30", "17:00", config=config)
        last_sig = result.iloc[-1]["signal"]
        last_close = result.iloc[-1]["Close"]
        if last_sig:
            print(f"  {end_time.strftime('%H:%M')} ({len(growing_df)} bars) -> SIGNAL: {last_sig} @ {last_close:.0f}")
        else:
            print(f"  {end_time.strftime('%H:%M')} ({len(growing_df)} bars) -> no signal")
    except Exception as e:
        print(f"  {end_time.strftime('%H:%M')} ({len(growing_df)} bars) -> ERROR: {e}")

# Also run on the full session to compare
print("\n=== FULL SESSION RUN ON FRED'S LIVE DATA ===")
result_full = run_trading_algo_fast(minute_df, "2026-06-23", "09:30", "17:00", config=config)
trades = result_full[result_full["signal"].isin(["BUY", "SELL"])]
print(f"Total trades: {len(trades)}")
print(f"First 5 trades:")
for idx, row in trades.head(5).iterrows():
    sig = row["signal"]
    price = row["buy_price"] if sig == "BUY" else row["sell_price"]
    print(f"  {idx.strftime('%H:%M')} {sig} @ {price:.0f}")
print(f"Final P/L: {result_full['session_pl'].iloc[-1]:.0f}")

# Compare with CSV data
print("\n=== FULL SESSION RUN ON CSV DATA ===")
csv_path = os.path.expanduser("~/Desktop/2YearsData/full_day/CBOT_MINI_YM1_2026-06-23.csv")
csv_df = pd.read_csv(csv_path, index_col=0, parse_dates=True)
csv_df.index = pd.to_datetime(csv_df.index, utc=True).tz_convert(EST)
csv_df = csv_df[(csv_df.index >= session_start) & (csv_df.index <= pd.Timestamp("2026-06-23 16:59", tz=EST))]
result_csv = run_trading_algo_fast(csv_df, "2026-06-23", "09:30", "17:00", config=config)
trades_csv = result_csv[result_csv["signal"].isin(["BUY", "SELL"])]
print(f"Total trades: {len(trades_csv)}")
print(f"First 5 trades:")
for idx, row in trades_csv.head(5).iterrows():
    sig = row["signal"]
    price = row["buy_price"] if sig == "BUY" else row["sell_price"]
    print(f"  {idx.strftime('%H:%M')} {sig} @ {price:.0f}")
print(f"Final P/L: {result_csv['session_pl'].iloc[-1]:.0f}")

# Show OHLC differences in first few bars
print("\n=== BAR DIFFERENCES (Live vs CSV) ===")
print(f"{'Time':<6} {'Live Close':>11} {'CSV Close':>10} {'Diff':>6}")
for i in range(min(10, len(minute_df), len(csv_df))):
    lt = minute_df.index[i]
    ct = csv_df.index[i]
    if lt == ct:
        lc = minute_df.iloc[i]["Close"]
        cc = csv_df.iloc[i]["Close"]
        diff = lc - cc
        flag = " ***" if abs(diff) > 1 else ""
        print(f"{lt.strftime('%H:%M'):<6} {lc:>11.0f} {cc:>10.0f} {diff:>+6.0f}{flag}")

